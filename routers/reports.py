"""Reports module — Financial reports + MVP catalog, filters, Excel/PDF export."""
from datetime import datetime, timedelta
from io import BytesIO
from pydantic import BaseModel, Field
from typing import Any, Dict, List, Literal, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

from core import db, get_current_user, is_super_admin, is_sports_admin, now_utc, resolve_user_institution, fee_entity_filter, format_date_display, format_datetime_display
from reports_engine import (
    _access_reports,
    resolve_entity,
    REPORT_CATALOG,
    RUNNERS,
    export_excel,
    export_pdf,
    dict_rows_to_matrix,
    _subtitle,
)

router = APIRouter(prefix="/reports", tags=["reports"])


# ---------------- helpers ----------------
FEE_HEADS = ["Registration", "Monthly", "Hostel", "Day Boarding", "Transport", "Uniform", "Kit", "Tournament", "Books", "Event", "Other"]

def _access_check(user: dict, area: str = "financial"):
    """Guard: super_admin, Sports Admin, or PWS principal/VP for financial reports."""
    role = user.get("role")
    if role == "super_admin":
        return
    if role == "admin":
        return
    if role in ("principal", "vice_principal"):
        return
    raise HTTPException(403, "You do not have access to Reports.")


def _resolve_institution(user: dict, requested: Optional[str]) -> str:
    return resolve_user_institution(user, requested)


def _entity_filter(inst: str) -> Dict[str, Any]:
    return fee_entity_filter(inst)


def _parse_iso(d: Optional[str]) -> Optional[datetime]:
    if not d:
        return None
    try:
        return datetime.fromisoformat(d[:10])
    except Exception:
        return None


def _payments_entity_filter(inst: str) -> Dict[str, Any]:
    if inst == "PWS":
        return {"entity_id": "pws"}
    if inst == "ALPHA":
        return {"entity_id": "alpha"}
    return {}


def _payment_paid_day(p: dict) -> str:
    return (p.get("transaction_date") or p.get("created_at") or "")[:10]


async def _fetch_invoice_payment_receipts(
    inst: str,
    start: str,
    end: str,
    centre: Optional[str] = None,
    limit: int = 5000,
) -> List[dict]:
    """Load invoice payments in range, optionally filtered by payer venue."""
    pay_q: Dict[str, Any] = {**_payments_entity_filter(inst), "status": {"$ne": "refunded"}}
    pay_q["$or"] = [
        {"transaction_date": {"$gte": start, "$lte": end}},
        {"created_at": {"$gte": f"{start}T00:00:00", "$lte": f"{end}T23:59:59"}},
    ]
    payments = await db.payments.find(pay_q, {"_id": 0}).sort("created_at", -1).to_list(limit)
    if not payments:
        return []

    inv_ids = list({p.get("invoice_id") for p in payments if p.get("invoice_id")})
    inv_map: Dict[str, dict] = {}
    if inv_ids:
        invs = await db.invoices.find({"id": {"$in": inv_ids}}, {"_id": 0, "id": 1, "invoice_number": 1, "person_name": 1, "person_id": 1}).to_list(len(inv_ids))
        inv_map = {i["id"]: i for i in invs}

    person_ids = list({inv_map.get(p.get("invoice_id"), {}).get("person_id") for p in payments if inv_map.get(p.get("invoice_id"), {}).get("person_id")})
    person_map: Dict[str, dict] = {}
    if person_ids:
        people = await db.people.find({"id": {"$in": person_ids}}, {"_id": 0, "id": 1, "name": 1, "centre": 1, "group": 1}).to_list(len(person_ids))
        person_map = {p["id"]: p for p in people}

    rows: List[dict] = []
    for p in payments:
        day = _payment_paid_day(p)
        if day < start or day > end:
            continue
        inv = inv_map.get(p.get("invoice_id"), {})
        person = person_map.get(inv.get("person_id"), {})
        venue = person.get("centre") or person.get("group") or ""
        if centre and centre != "All" and venue and venue != centre:
            continue
        rows.append({
            "id": p.get("id"),
            "source": "invoice_payment",
            "receipt_number": p.get("receipt_number"),
            "invoice_number": inv.get("invoice_number"),
            "payer_name": inv.get("person_name") or person.get("name"),
            "amount": int(p.get("amount") or 0),
            "payment_mode": p.get("payment_mode") or "Unknown",
            "reference_id": p.get("reference_id"),
            "paid_at": day,
            "paid_at_iso": p.get("created_at"),
            "venue": venue or None,
            "sport": None,
            "fee_type": "Invoice payment",
            "collected_by_name": p.get("collected_by_name"),
            "entity_id": p.get("entity_id"),
        })
    return rows


async def _fetch_legacy_fee_receipts(
    base_filter: Dict[str, Any],
    start: str,
    end: str,
    limit: int = 5000,
) -> List[dict]:
    """Load paid legacy fee receipts in range."""
    paid_q = {**base_filter, "status": "paid", "paid_at": {"$gte": start, "$lte": f"{end}T23:59:59"}}
    fees = await db.fees.find(paid_q, {"_id": 0}).sort("paid_at", -1).limit(limit).to_list(limit)
    rows: List[dict] = []
    for f in fees:
        paid_at = f.get("paid_at") or ""
        day = paid_at[:10]
        if not day or day < start or day > end:
            continue
        rows.append({
            "id": f.get("id") or f.get("batch_id"),
            "source": "legacy_fee",
            "receipt_number": f.get("receipt_number") or f.get("reference_id") or (f.get("batch_id") or f.get("id") or "")[:8],
            "invoice_number": None,
            "payer_name": f.get("player_name") or f.get("person_name"),
            "amount": int(f.get("amount_due") or 0),
            "payment_mode": f.get("payment_mode") or "Unknown",
            "reference_id": f.get("reference_id"),
            "paid_at": day,
            "paid_at_iso": paid_at,
            "venue": f.get("centre"),
            "sport": f.get("sport"),
            "fee_type": f.get("fee_type"),
            "collected_by_name": f.get("collected_by_name") or f.get("paid_by_name"),
            "entity_id": f.get("entity_id") or "alpha",
        })
    return rows


def _build_fee_query(
    user: dict,
    date_from: Optional[str],
    date_to: Optional[str],
    institution: Optional[str],
    centre: Optional[str],
    sport: Optional[str],
    status: Optional[str],
    payment_status: Optional[str],
) -> Dict[str, Any]:
    """Build MongoDB query for `fees` collection with the global filters applied."""
    q: Dict[str, Any] = {}
    inst = _resolve_institution(user, institution)
    q.update(_entity_filter(inst))
    if inst == "ALPHA":
        if centre and centre != "All":
            q["centre"] = centre
        if sport and sport != "All":
            q["sport"] = sport
    elif inst == "PWS":
        if centre and centre != "All":
            q["centre"] = centre
    if payment_status == "paid":
        q["status"] = "paid"
    elif payment_status == "pending":
        q["status"] = "due"
    elif payment_status == "overdue":
        # Overdue = due AND due_date < today
        q["status"] = "due"
        q["due_date"] = {"$lt": now_utc().strftime("%Y-%m-%d")}
    # Date range on paid_at for "collected within window"; on due_date if payment_status pending/overdue
    df = _parse_iso(date_from)
    dt = _parse_iso(date_to)
    if df or dt:
        # If filtering paid, use paid_at range; otherwise use due_date range
        field = "paid_at" if q.get("status") == "paid" else "due_date"
        rng: Dict[str, Any] = q.get(field, {}) if isinstance(q.get(field), dict) else {}
        if df:
            rng["$gte"] = df.strftime("%Y-%m-%d")
        if dt:
            rng["$lte"] = (dt.replace(hour=23, minute=59)).strftime("%Y-%m-%d")
        q[field] = rng
    return q


# ---------------- 1. Revenue & Fee Summary ----------------
async def _ensure_recurring_fees():
    """Materialize all players' recurring monthly dues before financial aggregation."""
    from routers.fees import ensure_all_players_monthly_fees
    await ensure_all_players_monthly_fees()


@router.get("/financial/summary")
async def revenue_summary(
    user: dict = Depends(get_current_user),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    institution: Optional[str] = None,
    centre: Optional[str] = None,
    sport: Optional[str] = None,
    status: Optional[str] = None,
):
    """Revenue & Fee Summary.

    Returns totals, current-month vs prev-month collections, outstanding dues,
    and breakdown by fee head. Drill-down data: Institution -> Branch -> Sport.
    """
    _access_check(user)
    await _ensure_recurring_fees()
    inst = _resolve_institution(user, institution)
    now = now_utc()
    cur_month = now.strftime("%Y-%m")
    prev_month_dt = (now.replace(day=1) - timedelta(days=1))
    prev_month = prev_month_dt.strftime("%Y-%m")

    base_filter = _entity_filter(inst)
    if centre and centre != "All":
        base_filter["centre"] = centre
    if sport and sport != "All" and inst != "PWS":
        base_filter["sport"] = sport

    df = _parse_iso(date_from)
    dt = _parse_iso(date_to)

    paid_filter = {**base_filter, "status": "paid"}
    if df or dt:
        rng: Dict[str, Any] = {}
        if df: rng["$gte"] = df.strftime("%Y-%m-%d")
        if dt: rng["$lte"] = dt.strftime("%Y-%m-%d")
        paid_filter["paid_at"] = rng

    due_filter = {**base_filter, "status": "due"}

    # Totals
    async def _sum(q, field="amount_due"):
        pipeline = [{"$match": q}, {"$group": {"_id": None, "sum": {"$sum": f"${field}"}}}]
        cur = db.fees.aggregate(pipeline)
        docs = await cur.to_list(1)
        return int(docs[0]["sum"]) if docs else 0

    collected_all = await _sum(paid_filter)
    outstanding = await _sum(due_filter)

    current_month_col = await _sum({**base_filter, "status": "paid", "paid_at": {"$gte": f"{cur_month}-01"}})
    previous_month_col = await _sum({
        **base_filter, "status": "paid",
        "paid_at": {"$gte": f"{prev_month}-01", "$lt": f"{cur_month}-01"}
    })

    # By fee head (respects filters)
    q_head = {**base_filter, "status": "paid"}
    if df or dt:
        q_head["paid_at"] = paid_filter["paid_at"]
    pipeline_head = [
        {"$match": q_head},
        {"$group": {"_id": "$fee_type", "sum": {"$sum": "$amount_due"}, "count": {"$sum": 1}}},
        {"$sort": {"sum": -1}},
    ]
    by_head_docs = await db.fees.aggregate(pipeline_head).to_list(50)
    by_head = [{"fee_head": d["_id"] or "Other", "amount": int(d["sum"]), "count": d["count"]} for d in by_head_docs]

    # Drill-down: by centre
    pipeline_c = [
        {"$match": q_head},
        {"$group": {"_id": "$centre", "collected": {"$sum": "$amount_due"}, "count": {"$sum": 1}}},
    ]
    by_centre_docs = await db.fees.aggregate(pipeline_c).to_list(20)
    by_centre = [{"centre": d["_id"] or "Unknown", "collected": int(d["collected"]), "count": d["count"]} for d in by_centre_docs]

    # Drill-down: by sport
    pipeline_s = [
        {"$match": q_head},
        {"$group": {"_id": "$sport", "collected": {"$sum": "$amount_due"}, "count": {"$sum": 1}}},
    ]
    by_sport_docs = await db.fees.aggregate(pipeline_s).to_list(20)
    by_sport = [{"sport": d["_id"] or "Unknown", "collected": int(d["collected"]), "count": d["count"]} for d in by_sport_docs]

    by_institution = []
    if inst in ("ALPHA", "BOTH"):
        alpha_filter = {**base_filter, **_entity_filter("ALPHA")}
        alpha_collected = await _sum({**alpha_filter, "status": "paid"})
        alpha_outstanding = await _sum({**alpha_filter, "status": "due"})
        by_institution.append({"institution": "ALPHA", "collected": alpha_collected, "outstanding": alpha_outstanding})
    if inst in ("PWS", "BOTH"):
        pws_filter = {**base_filter, **_entity_filter("PWS")}
        pws_collected = await _sum({**pws_filter, "status": "paid"})
        pws_outstanding = await _sum({**pws_filter, "status": "due"})
        by_institution.append({"institution": "PWS", "collected": pws_collected, "outstanding": pws_outstanding})
    return {
        "totals": {
            "collected_all_time": collected_all,
            "current_month": current_month_col,
            "previous_month": previous_month_col,
            "outstanding": outstanding,
        },
        "by_fee_head": by_head,
        "by_centre": by_centre,
        "by_sport": by_sport,
        "by_institution": by_institution,
    }


# ---------------- 2. Defaulter & Aging Report ----------------
@router.get("/financial/defaulters")
async def defaulters_aging(
    user: dict = Depends(get_current_user),
    institution: Optional[str] = None,
    centre: Optional[str] = None,
    sport: Optional[str] = None,
    limit: int = Query(500, le=2000),
):
    _access_check(user)
    await _ensure_recurring_fees()
    inst = _resolve_institution(user, institution)
    today_str = now_utc().strftime("%Y-%m-%d")
    q: Dict[str, Any] = {"status": "due", "due_date": {"$lt": today_str}}
    q.update(_entity_filter(inst))
    if centre and centre != "All":
        q["centre"] = centre
    if sport and sport != "All" and inst != "PWS":
        q["sport"] = sport
    cur = db.fees.find(q, {"_id": 0}).limit(limit)
    fees_docs = await cur.to_list(limit)
    rows = []
    today_dt = now_utc()
    buckets = {"0_7": 0, "8_15": 0, "16_30": 0, "gt_30": 0}
    for f in fees_docs:
        try:
            due_dt = datetime.fromisoformat(f.get("due_date", "")[:10])
            days = max((today_dt - due_dt).days, 0)
        except Exception:
            days = 0
        if days <= 7: bucket = "0_7"
        elif days <= 15: bucket = "8_15"
        elif days <= 30: bucket = "16_30"
        else: bucket = "gt_30"
        buckets[bucket] += 1
        rows.append({
            "player_id": f.get("player_id"),
            "player_name": f.get("player_name"),
            "centre": f.get("centre"),
            "sport": f.get("sport"),
            "category": f.get("category") or f.get("player_type"),
            "fee_type": f.get("fee_type"),
            "amount_due": int(f.get("amount_due", 0)),
            "due_date": f.get("due_date"),
            "days_overdue": days,
            "bucket": bucket,
        })
    rows.sort(key=lambda r: r["days_overdue"], reverse=True)
    return {"buckets": buckets, "rows": rows}


# ---------------- 3. Payment Mode Report ----------------
@router.get("/financial/payment-modes")
async def payment_modes(
    user: dict = Depends(get_current_user),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    institution: Optional[str] = None,
    centre: Optional[str] = None,
    sport: Optional[str] = None,
    limit: int = Query(1000, le=5000),
):
    _access_check(user)
    inst = _resolve_institution(user, institution)
    q: Dict[str, Any] = {"status": "paid"}
    q.update(_entity_filter(inst))
    if centre and centre != "All":
        q["centre"] = centre
    if sport and sport != "All" and inst != "PWS":
        q["sport"] = sport
    df = _parse_iso(date_from)
    dt = _parse_iso(date_to)
    if df or dt:
        rng: Dict[str, Any] = {}
        if df: rng["$gte"] = df.strftime("%Y-%m-%d")
        if dt: rng["$lte"] = dt.strftime("%Y-%m-%d")
        q["paid_at"] = rng
    # Summary aggregation
    pipeline = [
        {"$match": q},
        {"$group": {"_id": {"$ifNull": ["$payment_mode", "Unknown"]}, "count": {"$sum": 1}, "sum": {"$sum": "$amount_due"}}},
        {"$sort": {"sum": -1}},
    ]
    grp_docs = await db.fees.aggregate(pipeline).to_list(20)
    summary = {(d["_id"] or "Unknown"): {"count": d["count"], "sum": int(d["sum"])} for d in grp_docs}
    # Transactions (for online payments show ref/date/by)
    txn_cur = db.fees.find(q, {"_id": 0}).sort("paid_at", -1).limit(limit)
    txns = []
    async for f in txn_cur:
        txns.append({
            "player_id": f.get("player_id"),
            "player_name": f.get("player_name"),
            "centre": f.get("centre"),
            "sport": f.get("sport"),
            "fee_type": f.get("fee_type"),
            "amount": int(f.get("amount_due", 0)),
            "payment_mode": f.get("payment_mode") or "Unknown",
            "reference_id": f.get("reference_id"),
            "paid_at": f.get("paid_at"),
            "collected_by_name": f.get("collected_by_name") or f.get("paid_by_name") or None,
            "source": "legacy_fee",
        })

    inst = _resolve_institution(user, institution)
    start = (df or now_utc().replace(day=1)).strftime("%Y-%m-%d")
    end = (dt or now_utc()).strftime("%Y-%m-%d")
    inv_receipts = await _fetch_invoice_payment_receipts(inst, start, end, centre, limit)
    for r in inv_receipts:
        mode = r.get("payment_mode") or "Unknown"
        agg = summary.setdefault(mode, {"count": 0, "sum": 0})
        agg["count"] += 1
        agg["sum"] += r["amount"]
        txns.append({
            "player_id": None,
            "player_name": r.get("payer_name"),
            "centre": r.get("venue"),
            "sport": None,
            "fee_type": r.get("fee_type"),
            "amount": r["amount"],
            "payment_mode": mode,
            "reference_id": r.get("reference_id"),
            "paid_at": r.get("paid_at_iso") or r.get("paid_at"),
            "collected_by_name": r.get("collected_by_name"),
            "source": "invoice_payment",
            "receipt_number": r.get("receipt_number"),
            "invoice_number": r.get("invoice_number"),
        })
    txns.sort(key=lambda t: t.get("paid_at") or "", reverse=True)
    if len(txns) > limit:
        txns = txns[:limit]

    return {"summary": summary, "transactions": txns}


# ---------------- 3b. Daily Sales & Revenue Log ----------------
@router.get("/financial/daily-revenue-log")
async def daily_revenue_log(
    user: dict = Depends(get_current_user),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    institution: Optional[str] = None,
    centre: Optional[str] = None,
    sport: Optional[str] = None,
):
    """Daily collected revenue and expected dues for Finance Reports."""
    _access_check(user)
    await _ensure_recurring_fees()
    inst = _resolve_institution(user, institution)
    base_filter = _entity_filter(inst)
    if centre and centre != "All":
        base_filter["centre"] = centre
    if sport and sport != "All" and inst != "PWS":
        base_filter["sport"] = sport

    df = _parse_iso(date_from)
    dt = _parse_iso(date_to)
    now = now_utc()
    if not df:
        df = now.replace(day=1)
    if not dt:
        dt = now
    start = df.strftime("%Y-%m-%d")
    end = dt.strftime("%Y-%m-%d")

    paid_q = {**base_filter, "status": "paid", "paid_at": {"$gte": start, "$lte": end}}
    daily_collected = await db.fees.aggregate([
        {"$match": paid_q},
        {"$project": {"day": {"$substr": ["$paid_at", 0, 10]}, "amount_due": 1}},
        {"$group": {"_id": "$day", "collected": {"$sum": "$amount_due"}, "transactions": {"$sum": 1}}},
        {"$sort": {"_id": 1}},
    ]).to_list(400)
    collected_by_day = {d["_id"]: {"collected": int(d["collected"]), "transactions": d["transactions"]} for d in daily_collected}

    inv_receipts = await _fetch_invoice_payment_receipts(inst, start, end, centre)
    for r in inv_receipts:
        day = r["paid_at"]
        bucket = collected_by_day.setdefault(day, {"collected": 0, "transactions": 0})
        bucket["collected"] += r["amount"]
        bucket["transactions"] += 1

    expected_q = {**base_filter, "due_date": {"$gte": start, "$lte": end}}
    daily_expected = await db.fees.aggregate([
        {"$match": expected_q},
        {"$project": {"day": {"$substr": ["$due_date", 0, 10]}, "amount_due": 1}},
        {"$group": {"_id": "$day", "expected": {"$sum": "$amount_due"}}},
        {"$sort": {"_id": 1}},
    ]).to_list(400)
    expected_by_day = {d["_id"]: int(d["expected"]) for d in daily_expected}

    daily = []
    cursor = df
    while cursor.date() <= dt.date():
        day = cursor.strftime("%Y-%m-%d")
        c = collected_by_day.get(day, {"collected": 0, "transactions": 0})
        daily.append({
            "date": day,
            "collected": c["collected"],
            "transactions": c["transactions"],
            "expected": expected_by_day.get(day, 0),
        })
        cursor += timedelta(days=1)

    weekly: Dict[str, Dict[str, int]] = {}
    for row in daily:
        day_dt = datetime.fromisoformat(row["date"])
        week_start = day_dt - timedelta(days=day_dt.weekday())
        label = f"Week of {week_start.strftime('%d %b')}"
        bucket = weekly.setdefault(label, {"expected": 0, "collected": 0})
        bucket["expected"] += row["expected"]
        bucket["collected"] += row["collected"]

    trend = [{"label": k, **v} for k, v in weekly.items()]

    return {
        "period": {"from": start, "to": end},
        "daily": daily,
        "trend": trend,
        "totals": {
            "collected": sum(r["collected"] for r in daily),
            "expected": sum(r["expected"] for r in daily),
            "transactions": sum(r["transactions"] for r in daily),
        },
    }


# ---------------- 3c. Daily Collections & Receipts Log (unified) ----------------
@router.get("/financial/daily-collections")
async def daily_collections_log(
    user: dict = Depends(get_current_user),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    institution: Optional[str] = None,
    centre: Optional[str] = None,
    sport: Optional[str] = None,
    limit: int = Query(3000, le=5000),
):
    """Unified daily receipts register: legacy fee collections + invoice payments."""
    _access_check(user)
    await _ensure_recurring_fees()
    inst = _resolve_institution(user, institution)
    base_filter = _entity_filter(inst)
    if centre and centre != "All":
        base_filter["centre"] = centre
    if sport and sport != "All" and inst != "PWS":
        base_filter["sport"] = sport

    df = _parse_iso(date_from)
    dt = _parse_iso(date_to)
    now = now_utc()
    if not df:
        df = now.replace(day=1)
    if not dt:
        dt = now
    start = df.strftime("%Y-%m-%d")
    end = dt.strftime("%Y-%m-%d")

    legacy = await _fetch_legacy_fee_receipts(base_filter, start, end, limit)
    invoice = await _fetch_invoice_payment_receipts(inst, start, end, centre, limit)
    receipts = legacy + invoice
    receipts.sort(key=lambda r: (r.get("paid_at") or "", r.get("paid_at_iso") or ""), reverse=True)
    if len(receipts) > limit:
        receipts = receipts[:limit]

    by_day: Dict[str, Dict[str, int]] = {}
    for r in receipts:
        day = r.get("paid_at") or ""
        bucket = by_day.setdefault(day, {"count": 0, "amount": 0})
        bucket["count"] += 1
        bucket["amount"] += r["amount"]

    today = now.strftime("%Y-%m-%d")
    today_bucket = by_day.get(today, {"count": 0, "amount": 0})

    return {
        "period": {"from": start, "to": end},
        "receipts": receipts,
        "by_day": by_day,
        "totals": {
            "count": len(receipts),
            "amount": sum(r["amount"] for r in receipts),
            "collected_today": today_bucket["amount"],
            "transactions_today": today_bucket["count"],
        },
    }


# ---------------- 4. Excel Export ----------------
@router.get("/financial/export")
async def export_financial(
    kind: str = Query(..., description="One of: summary, defaulters, payment-modes"),
    user: dict = Depends(get_current_user),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    institution: Optional[str] = None,
    centre: Optional[str] = None,
    sport: Optional[str] = None,
    status: Optional[str] = None,
):
    _access_check(user)
    await _ensure_recurring_fees()
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E40AF")
    title_font = Font(bold=True, size=14, color="1E40AF")

    def write_title(sheet, title: str, subtitle: str):
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        c = sheet.cell(row=1, column=1, value=title)
        c.font = title_font
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        s = sheet.cell(row=2, column=1, value=subtitle)
        s.font = Font(italic=True, color="475569", size=10)

    def write_header(sheet, row: int, cols: List[str]):
        for idx, col in enumerate(cols, start=1):
            cell = sheet.cell(row=row, column=idx, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")

    subtitle = f"Filters — Institution: {_resolve_institution(user, institution)} · Centre: {centre or 'All'} · Sport: {sport or 'All'} · Range: {format_date_display(date_from) if date_from else '—'} → {format_date_display(date_to) if date_to else '—'} · Generated: {format_datetime_display(now_utc().isoformat())} by {user.get('name')}"

    if kind == "summary":
        ws.title = "Revenue Summary"
        data = await revenue_summary(user, date_from, date_to, institution, centre, sport, status)
        write_title(ws, "PWS & ALPHA — Revenue & Fee Summary", subtitle)
        # Totals block
        ws.cell(row=4, column=1, value="Total Collected (₹)").font = Font(bold=True)
        ws.cell(row=4, column=2, value=data["totals"]["collected_all_time"])
        ws.cell(row=5, column=1, value="Current Month (₹)").font = Font(bold=True)
        ws.cell(row=5, column=2, value=data["totals"]["current_month"])
        ws.cell(row=6, column=1, value="Previous Month (₹)").font = Font(bold=True)
        ws.cell(row=6, column=2, value=data["totals"]["previous_month"])
        ws.cell(row=7, column=1, value="Outstanding (₹)").font = Font(bold=True)
        ws.cell(row=7, column=2, value=data["totals"]["outstanding"])
        # By fee head
        write_header(ws, 9, ["Fee Head", "Amount (₹)", "Count"])
        for i, r in enumerate(data["by_fee_head"], start=10):
            ws.cell(row=i, column=1, value=r["fee_head"])
            ws.cell(row=i, column=2, value=r["amount"])
            ws.cell(row=i, column=3, value=r["count"])
        # By centre
        r_off = 10 + len(data["by_fee_head"]) + 2
        write_header(ws, r_off, ["Centre", "Collected (₹)", "Count"])
        for i, r in enumerate(data["by_centre"], start=r_off + 1):
            ws.cell(row=i, column=1, value=r["centre"])
            ws.cell(row=i, column=2, value=r["collected"])
            ws.cell(row=i, column=3, value=r["count"])
    elif kind == "defaulters":
        ws.title = "Defaulters & Aging"
        data = await defaulters_aging(user, institution, centre, sport, 2000)
        write_title(ws, "PWS & ALPHA — Defaulters & Aging", subtitle)
        b = data["buckets"]
        ws.cell(row=4, column=1, value="0–7 days").font = Font(bold=True); ws.cell(row=4, column=2, value=b["0_7"])
        ws.cell(row=5, column=1, value="8–15 days").font = Font(bold=True); ws.cell(row=5, column=2, value=b["8_15"])
        ws.cell(row=6, column=1, value="16–30 days").font = Font(bold=True); ws.cell(row=6, column=2, value=b["16_30"])
        ws.cell(row=7, column=1, value="More than 30 days").font = Font(bold=True); ws.cell(row=7, column=2, value=b["gt_30"])
        write_header(ws, 9, ["Player", "Centre", "Sport", "Category", "Fee Head", "Amount Due (₹)", "Due Date", "Days Overdue", "Bucket"])
        for i, r in enumerate(data["rows"], start=10):
            ws.cell(row=i, column=1, value=r.get("player_name"))
            ws.cell(row=i, column=2, value=r.get("centre"))
            ws.cell(row=i, column=3, value=r.get("sport"))
            ws.cell(row=i, column=4, value=r.get("category"))
            ws.cell(row=i, column=5, value=r.get("fee_type"))
            ws.cell(row=i, column=6, value=r.get("amount_due"))
            ws.cell(row=i, column=7, value=format_date_display(r.get("due_date")))
            ws.cell(row=i, column=8, value=r.get("days_overdue"))
            ws.cell(row=i, column=9, value=r.get("bucket"))
    elif kind == "payment-modes":
        ws.title = "Payment Modes"
        data = await payment_modes(user, date_from, date_to, institution, centre, sport, 5000)
        write_title(ws, "PWS & ALPHA — Payment Mode Report", subtitle)
        write_header(ws, 4, ["Mode", "Transactions", "Amount (₹)"])
        row = 5
        for mode, agg in data["summary"].items():
            ws.cell(row=row, column=1, value=mode)
            ws.cell(row=row, column=2, value=agg["count"])
            ws.cell(row=row, column=3, value=agg["sum"])
            row += 1
        row += 2
        write_header(ws, row, ["Player", "Centre", "Sport", "Fee Head", "Amount (₹)", "Mode", "Reference", "Paid At", "Collected By"])
        for i, t in enumerate(data["transactions"], start=row + 1):
            ws.cell(row=i, column=1, value=t.get("player_name"))
            ws.cell(row=i, column=2, value=t.get("centre"))
            ws.cell(row=i, column=3, value=t.get("sport"))
            ws.cell(row=i, column=4, value=t.get("fee_type"))
            ws.cell(row=i, column=5, value=t.get("amount"))
            ws.cell(row=i, column=6, value=t.get("payment_mode"))
            ws.cell(row=i, column=7, value=t.get("reference_id"))
            ws.cell(row=i, column=8, value=format_datetime_display(t.get("paid_at")))
            ws.cell(row=i, column=9, value=t.get("collected_by_name"))
    else:
        raise HTTPException(400, f"Unknown export kind: {kind}. Use one of: summary, defaulters, payment-modes")

    # Auto-widen columns
    for col_cells in ws.columns:
        try:
            max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)
        except Exception:
            pass

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"pws-alpha-{kind}-{now_utc().strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


class FinanceReportExportIn(BaseModel):
    organization: str = "ALPHA Sports Academy & PWS"
    title: str
    subtitle: str
    report_view: str
    format: Literal["csv", "xlsx", "pdf"]
    columns: List[str]
    rows: List[List[Any]]
    summary_rows: List[List[Any]] = Field(default_factory=list)


def _finance_export_pdf(org: str, title: str, subtitle: str, columns: List[str], rows: List[List[Any]], summary_rows: List[List[Any]], filename: str) -> StreamingResponse:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas as pdfcanvas

    page_size = landscape(A4)
    W, H = page_size
    margin = 20 * mm
    line_h = 5 * mm
    col_w = (W - 2 * margin) / max(len(columns), 1)

    def render_pages(c, footer_fn):
        page_num = 1
        y = H - margin

        def new_page():
            nonlocal page_num, y
            footer_fn(page_num)
            c.showPage()
            page_num += 1
            y = H - margin
            c.setFont("Helvetica", 7)

        c.setFont("Helvetica-Bold", 12)
        c.drawString(margin, y, org)
        y -= line_h
        c.setFont("Helvetica-Bold", 11)
        c.drawString(margin, y, title)
        y -= line_h
        c.setFont("Helvetica", 8)
        c.drawString(margin, y, subtitle[:160])
        y -= line_h * 2
        c.setFont("Helvetica-Bold", 8)
        for i, col in enumerate(columns):
            c.drawString(margin + i * col_w, y, str(col)[:20])
        y -= line_h
        c.setFont("Helvetica", 7)

        for row in rows:
            if y < 18 * mm:
                new_page()
                c.setFont("Helvetica-Bold", 8)
                for i, col in enumerate(columns):
                    c.drawString(margin + i * col_w, y, str(col)[:20])
                y -= line_h
                c.setFont("Helvetica", 7)
            for i, val in enumerate(row):
                c.drawString(margin + i * col_w, y, str(val)[:24] if val is not None else "")
            y -= line_h

        if summary_rows:
            y -= line_h
            c.setFont("Helvetica-Bold", 8)
            c.drawString(margin, y, "Summary")
            y -= line_h
            c.setFont("Helvetica", 7)
            for srow in summary_rows:
                if y < 18 * mm:
                    new_page()
                c.drawString(margin, y, " · ".join(str(x) for x in srow))
                y -= line_h

        footer_fn(page_num)
        return page_num

    count_buf = BytesIO()
    total_pages = render_pages(pdfcanvas.Canvas(count_buf, pagesize=page_size), lambda _p: None)

    out_buf = BytesIO()
    out_canvas = pdfcanvas.Canvas(out_buf, pagesize=page_size)

    def draw_footer(page_num: int):
        out_canvas.setFont("Helvetica", 7)
        out_canvas.drawString(margin, 10 * mm, subtitle[:160])
        out_canvas.drawRightString(W - margin, 10 * mm, f"Page {page_num} of {total_pages}")

    render_pages(out_canvas, draw_footer)
    out_canvas.save()
    out_buf.seek(0)
    return StreamingResponse(
        out_buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/finance-reports/export")
async def export_finance_report(payload: FinanceReportExportIn, user: dict = Depends(get_current_user)):
    _access_check(user)
    stamp = now_utc().strftime("%Y%m%d-%H%M")
    fname_base = f"finance-{payload.report_view}-{stamp}"
    subtitle = payload.subtitle

    if payload.format == "csv":
        lines = [payload.organization, payload.title, subtitle, "", ",".join(payload.columns)]
        for row in payload.rows:
            lines.append(",".join(f'"{str(c).replace(chr(34), chr(34)+chr(34))}"' for c in row))
        if payload.summary_rows:
            lines.extend(["", "Summary"])
            for srow in payload.summary_rows:
                lines.append(",".join(str(x) for x in srow))
        content = "\n".join(lines)
        return StreamingResponse(
            iter([content]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{fname_base}.csv"'},
        )

    matrix = [[str(c) for c in row] for row in payload.rows]
    if payload.format == "pdf":
        return _finance_export_pdf(
            payload.organization,
            payload.title,
            subtitle,
            payload.columns,
            matrix,
            payload.summary_rows,
            f"{fname_base}.pdf",
        )

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = payload.title[:31]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(payload.columns), 1))
    ws.cell(row=1, column=1, value=payload.organization).font = Font(bold=True, size=12, color="1E40AF")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(payload.columns), 1))
    ws.cell(row=2, column=1, value=payload.title).font = Font(bold=True, size=14, color="1E40AF")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max(len(payload.columns), 1))
    ws.cell(row=3, column=1, value=subtitle).font = Font(italic=True, color="475569", size=10)
    hr = 5
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E40AF")
    for idx, col in enumerate(payload.columns, start=1):
        cell = ws.cell(row=hr, column=idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")
    for ri, row in enumerate(matrix, start=hr + 1):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=val)
    if payload.summary_rows:
        off = hr + len(matrix) + 2
        ws.cell(row=off, column=1, value="Summary").font = Font(bold=True)
        for i, srow in enumerate(payload.summary_rows, start=off + 1):
            ws.cell(row=i, column=1, value=" · ".join(str(x) for x in srow))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname_base}.xlsx"'},
    )


# ---------------- MVP Reports (catalog, JSON, Excel/PDF export) ----------------

def _report_filters(
    entity: Optional[str],
    date_from: Optional[str],
    date_to: Optional[str],
    grade: Optional[str],
    section_id: Optional[str],
    sport: Optional[str],
    centre: Optional[str],
    status: Optional[str],
    player_type: Optional[str] = None,
    fee_collection_type: Optional[str] = None,
    payment_method: Optional[str] = None,
    pws_student_type: Optional[str] = None,
    department: Optional[str] = None,
    designation: Optional[str] = None,
    employment_type: Optional[str] = None,
    shift: Optional[str] = None,
) -> dict:
    return {
        k: v for k, v in {
            "entity": entity,
            "date_from": date_from,
            "date_to": date_to,
            "grade": grade,
            "section_id": section_id,
            "sport": sport,
            "centre": centre,
            "status": status,
            "player_type": player_type,
            "fee_collection_type": fee_collection_type,
            "payment_method": payment_method,
            "pws_student_type": pws_student_type,
            "department": department,
            "designation": designation,
            "employment_type": employment_type,
            "shift": shift,
        }.items() if v
    }


@router.get("/catalog")
async def report_catalog(user: dict = Depends(get_current_user)):
    """List available MVP reports and supported filters."""
    _access_reports(user)
    inst = resolve_user_institution(user, None)
    entity_options = ["PWS", "ALPHA"] if inst == "BOTH" else [inst]
    if is_super_admin(user) or inst == "BOTH":
        entity_options = ["PWS", "ALPHA", "Combined"]
    return {
        "reports": REPORT_CATALOG,
        "entity_options": entity_options,
        "export_formats": ["xlsx", "pdf"],
    }


@router.get("/{report_id}/export")
async def export_mvp_report(
    report_id: str,
    user: dict = Depends(get_current_user),
    format: str = Query("xlsx", description="xlsx or pdf"),
    entity: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    grade: Optional[str] = None,
    section_id: Optional[str] = None,
    sport: Optional[str] = None,
    centre: Optional[str] = None,
    status: Optional[str] = None,
    player_type: Optional[str] = None,
    fee_collection_type: Optional[str] = None,
    payment_method: Optional[str] = None,
    pws_student_type: Optional[str] = None,
    department: Optional[str] = None,
    designation: Optional[str] = None,
    employment_type: Optional[str] = None,
    shift: Optional[str] = None,
):
    _access_reports(user)
    runner = RUNNERS.get(report_id)
    if not runner:
        raise HTTPException(404, f"Unknown report: {report_id}")
    inst = resolve_entity(user, entity)
    filters = _report_filters(
        entity, date_from, date_to, grade, section_id, sport, centre, status, player_type,
        fee_collection_type, payment_method, pws_student_type,
        department, designation, employment_type, shift,
    )
    meta = await runner(user, inst, filters)
    columns, matrix = dict_rows_to_matrix(meta)
    subtitle = _subtitle(inst, filters, user)
    fmt = (format or "xlsx").lower()
    stamp = now_utc().strftime("%Y%m%d-%H%M")
    fname = f"{report_id}-{stamp}.{fmt}"
    if fmt == "pdf":
        return export_pdf(meta["title"], columns, matrix, subtitle, fname)
    if fmt == "xlsx":
        return export_excel(meta["title"], columns, matrix, subtitle, fname)
    raise HTTPException(400, "format must be xlsx or pdf")


@router.get("/{report_id}")
async def run_mvp_report(
    report_id: str,
    user: dict = Depends(get_current_user),
    entity: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    grade: Optional[str] = None,
    section_id: Optional[str] = None,
    sport: Optional[str] = None,
    centre: Optional[str] = None,
    status: Optional[str] = None,
    player_type: Optional[str] = None,
    fee_collection_type: Optional[str] = None,
    payment_method: Optional[str] = None,
    pws_student_type: Optional[str] = None,
    department: Optional[str] = None,
    designation: Optional[str] = None,
    employment_type: Optional[str] = None,
    shift: Optional[str] = None,
):
    """Run an MVP report and return JSON (table rows + summary)."""
    _access_reports(user)
    runner = RUNNERS.get(report_id)
    if not runner:
        raise HTTPException(404, f"Unknown report: {report_id}")
    inst = resolve_entity(user, entity)
    filters = _report_filters(
        entity, date_from, date_to, grade, section_id, sport, centre, status, player_type,
        fee_collection_type, payment_method, pws_student_type,
        department, designation, employment_type, shift,
    )
    return await runner(user, inst, filters)
