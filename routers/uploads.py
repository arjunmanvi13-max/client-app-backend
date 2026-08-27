"""Bulk Upload — spec-driven ingestion of students, players, staff and teachers.

The workbook produced by GET /bulk-upload/template is the accepted upload format:
download, fill in, upload. Columns come from bulk_ingest.SPECS, which covers every
field the roster and directory forms expose.

Ingestion is all-or-nothing per request: every row is validated before anything is
written, and if a write fails part-way the rows already created in that request are
removed, so a failed upload leaves the database as it was.
"""
import csv
import io
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, HTTPException, Query, Request, Response, UploadFile, File
from fastapi.responses import PlainTextResponse
from starlette.concurrency import run_in_threadpool

from core import db, get_current_user, logger, notify_role, PersonCreate, DirectoryTeacherCreate
from rbac.guards import can_bulk_upload
from bulk_ingest import KINDS, SPECS, SPEC_BY_SHEET, SheetSpec, _normalize_header, duplicate_key, parse_row
from bulk_template import build_csv, build_workbook

router = APIRouter(prefix="/bulk-upload", tags=["bulk-upload"])

ROW_LIMIT = 500
MAX_UPLOAD_BYTES = 8 * 1024 * 1024
XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

PEOPLE_KINDS = ("student", "player", "staff")

CLASS_TO_GRADE = {
    "Nursery": ["Nursery", "Nur"],
    "LKG": ["LKG"],
    "UKG": ["UKG"],
    "Class I": ["1", "I"],
    "Class II": ["2", "II"],
    "Class III": ["3", "III"],
    "Class IV": ["4", "IV"],
    "Class V": ["5", "V"],
    "Class VI": ["6", "VI"],
    "Class VII": ["7", "VII"],
    "Class VIII": ["8", "VIII"],
    "Class IX": ["9", "IX"],
    "Class X": ["10", "X"],
}


def _assert_can_upload(user: dict) -> None:
    if not can_bulk_upload(user):
        raise HTTPException(403, "bulk_upload permission required")


def _assert_declared_size(request: Request) -> None:
    """Reject on Content-Length before the body is parsed into a spooled file."""
    declared = request.headers.get("content-length")
    if declared and declared.isdigit() and int(declared) > MAX_UPLOAD_BYTES + 64 * 1024:
        raise HTTPException(413, f"File exceeds the {MAX_UPLOAD_BYTES // (1024 * 1024)} MB limit")


async def _read_capped(file: UploadFile) -> bytes:
    """Read the upload with a running byte cap.

    `_assert_declared_size` is the real guard — by the time this runs, Starlette
    has already spooled the multipart body. This caps the in-memory copy.
    """
    chunks: List[bytes] = []
    total = 0
    while True:
        chunk = await file.read(256 * 1024)
        if not chunk:
            break
        total += len(chunk)
        if total > MAX_UPLOAD_BYTES:
            raise HTTPException(413, f"File exceeds the {MAX_UPLOAD_BYTES // (1024 * 1024)} MB limit")
        chunks.append(chunk)
    if not total:
        raise HTTPException(400, "File is empty")
    return b"".join(chunks)


def _strip_required_marker(header: Any) -> str:
    s = "" if header is None else str(header)
    return s.replace("*", "").strip()


def _parse_csv(raw: bytes, spec: Optional[SheetSpec] = None) -> List[Dict[str, str]]:
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [r for r in reader]
    if not rows:
        return []
    headers = [_strip_required_marker(h) for h in rows[0]]
    out: List[Dict[str, str]] = []
    for r in rows[1:]:
        if all((c or "").strip() == "" for c in r):
            continue
        if not out and spec is not None and _is_guidance_row(list(r), spec, headers):
            continue
        out.append(_row_dict(headers, list(r)))
    return out


def _row_dict(headers: List[str], values: List[Any]) -> Dict[str, str]:
    """Extra cells past the last header become synthetic columns so parse_row
    reports them as unrecognised instead of dropping them silently."""
    row = {
        headers[i]: ("" if (i >= len(values) or values[i] is None) else str(values[i]))
        for i in range(len(headers))
    }
    for i in range(len(headers), len(values)):
        if str(values[i] or "").strip():
            row[f"Column {i + 1}"] = str(values[i])
    return row


def _guidance_text(spec: SheetSpec) -> List[str]:
    """The row-2 text build_workbook writes, matched exactly to skip it."""
    return [(c.help or ("Required" if c.required else "")).strip().lower() for c in spec.columns]


def _is_guidance_row(values: List[Any], spec: SheetSpec, headers: List[str]) -> bool:
    """Row 2 of the template holds column guidance — skip it rather than parse it.

    Matching the generated text exactly is what keeps a genuine first data row
    that happens to contain a word like "required" from being swallowed. The
    marker fallback only fires when every filled cell looks like guidance.
    """
    cells = [str(v).strip().lower() if v is not None else "" for v in values]
    if not any(cells):
        return True

    expected = _guidance_text(spec)
    trimmed = cells[: len(expected)]
    trimmed += [""] * (len(expected) - len(trimmed))
    if trimmed == [e if e else "none" for e in expected] or trimmed == expected:
        return True
    if all(c in ("", "none", e) for c, e in zip(trimmed, expected)):
        return True

    markers = ("one of:", "required", "yes / no", "10-digit", "blank =", "free text")
    filled = [c for c in cells if c and c != "none"]
    return bool(filled) and all(any(m in c for m in markers) for c in filled)


def _assert_unique_headers(headers: List[str], sheet_name: str) -> None:
    seen, repeated = set(), []
    for h in headers:
        key = _normalize_header(h or "")
        if not key:
            continue
        if key in seen:
            repeated.append(h)
        seen.add(key)
    if repeated:
        raise HTTPException(
            400, f"{sheet_name}: duplicate column heading(s): {', '.join(sorted(set(repeated)))}"
        )


def _sheet_rows(ws, spec: SheetSpec) -> List[Dict[str, str]]:
    headers: List[str] = []
    out: List[Dict[str, str]] = []
    for index, raw in enumerate(ws.iter_rows(values_only=True)):
        if index == 0:
            headers = [_strip_required_marker(h) for h in raw]
            _assert_unique_headers(headers, spec.sheet_name)
            continue
        if len(out) > ROW_LIMIT:
            raise HTTPException(
                400, f"{spec.sheet_name}: row limit is {ROW_LIMIT} per sheet"
            )
        values = list(raw)
        if all(v is None or str(v).strip() == "" for v in values):
            continue
        if not out and _is_guidance_row(values, spec, headers):
            continue
        out.append(_row_dict(headers, values))
    return out


def _parse_workbook(raw: bytes) -> Dict[str, List[Dict[str, str]]]:
    """Return {kind: rows} for every recognised sheet in the workbook."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    found: Dict[str, List[Dict[str, str]]] = {}
    for ws in wb.worksheets:
        spec = SPEC_BY_SHEET.get("".join(ch for ch in ws.title.lower() if ch.isalnum()))
        if spec is None:
            continue
        rows = _sheet_rows(ws, spec)
        if rows:
            found[spec.kind] = rows
    return found


def _parse_single_sheet(raw: bytes, filename: str, spec: SheetSpec) -> List[Dict[str, str]]:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or raw[:2] == b"PK":
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        target = None
        for ws in wb.worksheets:
            key = "".join(ch for ch in ws.title.lower() if ch.isalnum())
            if SPEC_BY_SHEET.get(key) is spec:
                target = ws
                break
        if target is None:
            target = wb.worksheets[0]
        return _sheet_rows(target, spec)
    return _parse_csv(raw, spec)


async def _validate_rows(
    spec: SheetSpec, rows: List[Dict[str, str]]
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Validate every row. Returns (parsed_values, errors)."""
    parsed: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    seen: Dict[Tuple, int] = {}

    for offset, raw_row in enumerate(rows):
        sheet_row = offset + 3
        values, errs = parse_row(spec, raw_row)

        key = duplicate_key(spec, values)
        if key and not errs:
            if key in seen:
                errs.append(f"Duplicate of row {seen[key]} in this file (same name and date of birth)")
            else:
                seen[key] = sheet_row

        if errs:
            errors.append({
                "sheet": spec.sheet_name,
                "row": sheet_row,
                "name": (raw_row.get("Full Name") or raw_row.get("Name") or "").strip(),
                "errors": errs,
            })
        else:
            parsed.append({"row": sheet_row, "values": values})

    return parsed, errors


async def _resolve_section_id(pws_class: Optional[str], letter: Optional[str]) -> Optional[str]:
    if not pws_class or not letter:
        return None
    for grade_name in CLASS_TO_GRADE.get(pws_class, []):
        section = await db.sections.find_one(
            {"grade_name": grade_name, "name": letter}, {"_id": 0, "id": 1}
        )
        if section:
            return section["id"]
    label_candidates = [f"{g}-{letter}" for g in CLASS_TO_GRADE.get(pws_class, [])]
    if label_candidates:
        section = await db.sections.find_one(
            {"label": {"$in": label_candidates}}, {"_id": 0, "id": 1}
        )
        if section:
            return section["id"]
    return None


async def _existing_name_warning(values: Dict[str, Any]) -> Optional[str]:
    """A namesake with a different date of birth is legitimate — surface it, do not block."""
    name = (values.get("name") or "").strip()
    if not name:
        return None
    hit = await db.people.find_one(
        {"name": {"$regex": f"^{_escape(name)}$", "$options": "i"}},
        {"_id": 0, "id": 1, "kind": 1, "dob": 1},
    )
    if not hit:
        return None
    if values.get("dob") and hit.get("dob") == values.get("dob"):
        return None
    existing_dob = hit.get("dob") or "no date of birth on record"
    return (f"An existing {hit.get('kind', 'person')} is also called {name} "
            f"({existing_dob}). Created as a separate record — check this is not a duplicate.")


async def _existing_person_conflict(values: Dict[str, Any]) -> Optional[str]:
    """Reject rows that would duplicate a person already in the database."""
    name = (values.get("name") or "").strip()
    dob = values.get("dob")
    if name and dob:
        hit = await db.people.find_one(
            {"kind": {"$in": ["student", "player"]}, "dob": dob,
             "name": {"$regex": f"^{_escape(name)}$", "$options": "i"}},
            {"_id": 0, "id": 1},
        )
        if hit:
            return "A person with this name and date of birth already exists"
    for field, label in (("admission_number", "Admission Number"), ("employee_id", "Employee ID"),
                         ("player_id", "Player ID")):
        val = values.get(field)
        if val and await db.people.find_one({field: val}, {"_id": 0, "id": 1}):
            return f"{label} {val} is already in use"
    return None


def _escape(s: str) -> str:
    import re
    return re.escape(s)


async def _create_people(spec: SheetSpec, parsed: List[Dict[str, Any]], user: dict) -> Dict[str, Any]:
    from routers.people import create_person

    created_ids: List[str] = []
    approvals = 0
    failures: List[Dict[str, Any]] = []

    for entry in parsed:
        values = dict(entry["values"])
        letter = values.pop("_section_letter", None)
        values["kind"] = spec.kind
        try:
            if spec.kind == "student":
                values.setdefault("organization", "PWS")
                section_id = await _resolve_section_id(values.get("pws_class"), letter)
                if section_id:
                    values["section_id"] = section_id
            elif spec.kind == "player":
                values.setdefault("organization", "ALPHA")
                if values.get("pws_class"):
                    section_id = await _resolve_section_id(values.get("pws_class"), letter)
                    if section_id:
                        values["section_id"] = section_id
            payload = PersonCreate(**values)
        except Exception as exc:
            failures.append({"sheet": spec.sheet_name, "row": entry["row"],
                             "name": values.get("name", ""), "errors": [_pydantic_message(exc)]})
            break

        try:
            result = await create_person(payload=payload, user=user)
        except HTTPException as exc:
            failures.append({"sheet": spec.sheet_name, "row": entry["row"],
                             "name": values.get("name", ""), "errors": [str(exc.detail)]})
            break
        except Exception as exc:
            logger.exception("Bulk upload failed on %s row %s", spec.sheet_name, entry["row"])
            failures.append({"sheet": spec.sheet_name, "row": entry["row"],
                             "name": values.get("name", ""), "errors": [str(exc)]})
            break

        person = result.get("person") if isinstance(result, dict) and "person" in result else result
        if isinstance(person, dict) and person.get("id"):
            created_ids.append(person["id"])
        if isinstance(result, dict) and result.get("approval_required"):
            approvals += 1

    return {"created_ids": created_ids, "approvals": approvals, "failures": failures}


async def _create_teachers(parsed: List[Dict[str, Any]], user: dict) -> Dict[str, Any]:
    from routers.users import create_directory_teacher

    created_ids: List[str] = []
    failures: List[Dict[str, Any]] = []

    for entry in parsed:
        values = dict(entry["values"])
        values.pop("teacher_designation", None)
        values.pop("date_of_joining", None)
        values.setdefault("enable_login", False)
        if not values.get("enable_login"):
            values.pop("login_email", None)
        try:
            payload = DirectoryTeacherCreate(**values)
        except Exception as exc:
            failures.append({"sheet": "Teachers", "row": entry["row"],
                             "name": values.get("name", ""), "errors": [_pydantic_message(exc)]})
            break
        try:
            result = await create_directory_teacher(payload=payload, user=user)
        except HTTPException as exc:
            failures.append({"sheet": "Teachers", "row": entry["row"],
                             "name": values.get("name", ""), "errors": [str(exc.detail)]})
            break
        except Exception as exc:
            logger.exception("Bulk teacher upload failed on row %s", entry["row"])
            failures.append({"sheet": "Teachers", "row": entry["row"],
                             "name": values.get("name", ""), "errors": [str(exc)]})
            break
        if isinstance(result, dict) and result.get("id"):
            created_ids.append(result["id"])

    return {"created_ids": created_ids, "approvals": 0, "failures": failures}


def _pydantic_message(exc: Exception) -> str:
    errs = getattr(exc, "errors", None)
    if callable(errs):
        try:
            parts = []
            for e in errs():
                loc = ".".join(str(x) for x in e.get("loc", ()) if x != "body")
                parts.append(f"{loc}: {e.get('msg')}" if loc else str(e.get("msg")))
            return "; ".join(parts) or str(exc)
        except Exception:
            pass
    return str(exc)


async def _rollback(kind: str, created_ids: List[str]) -> List[str]:
    """Undo one kind's writes, children first. Returns ids that could not be cleaned."""
    if not created_ids:
        return []

    legs: List[Tuple[str, Any, dict]] = []
    if kind in ("student", "player"):
        legs.append(("fees", db.fees, {"player_id": {"$in": created_ids}}))
        legs.append(("approval_requests", db.approval_requests, {"subject_id": {"$in": created_ids}}))
    if kind == "staff":
        legs.append(("users", db.users, {"person_id": {"$in": created_ids}}))
    if kind == "teacher":
        legs.append(("users", db.users, {"id": {"$in": created_ids}}))
    else:
        legs.append(("people", db.people, {"id": {"$in": created_ids}}))

    stranded: List[str] = []
    for label, coll, query in legs:
        try:
            await coll.delete_many(query)
        except Exception:
            logger.exception(
                "Rollback leg %s failed for %s ids %s — manual cleanup required",
                label, kind, created_ids,
            )
            stranded.extend(created_ids)
    if not stranded:
        logger.warning("Rolled back %s partially-created %s record(s)", len(created_ids), kind)
    return stranded


async def _rollback_all(created: Dict[str, List[str]]) -> List[str]:
    stranded: List[str] = []
    for done_kind, ids in created.items():
        stranded.extend(await _rollback(done_kind, ids))
    return stranded


def _model_construction_error(spec: SheetSpec, values: Dict[str, Any]) -> Optional[str]:
    """Run the same pydantic model the write path uses, so a dry run cannot pass
    a row that the import would then reject and roll the whole batch back for."""
    candidate = dict(values)
    candidate.pop("_section_letter", None)
    try:
        if spec.kind == "teacher":
            DirectoryTeacherCreate(**candidate)
        else:
            candidate["kind"] = spec.kind
            candidate.setdefault(
                "organization", "PWS" if spec.kind == "student" else "ALPHA"
            )
            PersonCreate(**candidate)
    except Exception as exc:
        return _pydantic_message(exc)
    return None


async def _ingest(sheets: Dict[str, List[Dict[str, str]]], user: dict, *, dry_run: bool) -> Dict[str, Any]:
    total_rows = sum(len(r) for r in sheets.values())
    if total_rows == 0:
        raise HTTPException(400, "No data rows found. Fill in the template starting at row 3.")
    if total_rows > ROW_LIMIT:
        raise HTTPException(400, f"Row limit is {ROW_LIMIT} per upload (found {total_rows})")

    validated: Dict[str, List[Dict[str, Any]]] = {}
    errors: List[Dict[str, Any]] = []
    warnings: List[Dict[str, Any]] = []

    for kind, rows in sheets.items():
        spec = SPECS[kind]
        parsed, errs = await _validate_rows(spec, rows)
        errors.extend(errs)
        keep: List[Dict[str, Any]] = []
        for entry in parsed:
            conflict = await _existing_person_conflict(entry["values"]) if kind in PEOPLE_KINDS else None
            if conflict:
                errors.append({"sheet": spec.sheet_name, "row": entry["row"],
                               "name": entry["values"].get("name", ""), "errors": [conflict]})
                continue
            model_error = _model_construction_error(spec, entry["values"])
            if model_error:
                errors.append({"sheet": spec.sheet_name, "row": entry["row"],
                               "name": entry["values"].get("name", ""), "errors": [model_error]})
                continue
            if kind in PEOPLE_KINDS:
                note = await _existing_name_warning(entry["values"])
                if note:
                    warnings.append({"sheet": spec.sheet_name, "row": entry["row"],
                                     "name": entry["values"].get("name", ""), "warning": note})
            keep.append(entry)
        validated[kind] = keep

    counts = {k: len(v) for k, v in validated.items() if v}

    if errors:
        return {
            "status": "validation_failed",
            "valid_count": sum(counts.values()),
            "counts": counts,
            "errors": sorted(errors, key=lambda e: (e["sheet"], e["row"])),
            "warnings": warnings,
        }

    if dry_run:
        return {"status": "ok", "dry_run": True, "counts": counts,
                "valid_count": sum(counts.values()), "errors": [], "warnings": warnings}

    created: Dict[str, List[str]] = {}
    approvals = 0
    try:
        for kind, entries in validated.items():
            if not entries:
                continue
            if kind == "teacher":
                outcome = await _create_teachers(entries, user)
            else:
                outcome = await _create_people(SPECS[kind], entries, user)
            created[kind] = outcome["created_ids"]
            approvals += outcome["approvals"]
            if outcome["failures"]:
                stranded = await _rollback_all(created)
                detail = "Upload rolled back — nothing was saved."
                if stranded:
                    detail = (
                        "Upload failed and could not be fully rolled back. "
                        f"These records need manual cleanup: {', '.join(stranded)}"
                    )
                return {
                    "status": "failed",
                    "created": {},
                    "counts": {},
                    "errors": outcome["failures"],
                    "detail": detail,
                }
    except BaseException:
        await _rollback_all(created)
        raise

    created_counts = {k: len(v) for k, v in created.items()}
    total_created = sum(created_counts.values())
    if total_created:
        summary = ", ".join(f"{v} {k}" for k, v in created_counts.items() if v)
        await notify_role(
            "super_admin",
            ntype="bulk_upload_completed",
            title="Bulk upload completed",
            message=f"{user['name']} uploaded {summary}"
                    + (f" · {approvals} awaiting fee approval" if approvals else ""),
        )
    billable_ids = [pid for kind in ("student", "player") for pid in created.get(kind, [])]
    fees_created = (
        await db.fees.count_documents({"player_id": {"$in": billable_ids}}) if billable_ids else 0
    )
    if billable_ids and not fees_created:
        warnings.append({
            "sheet": "-", "row": 0, "name": "",
            "errors": ["Records were imported but no fee schedule was generated — check the fee catalogue."],
        })
    return {
        "status": "ok",
        "counts": created_counts,
        "created_total": total_created,
        "fees_created": fees_created,
        "pending_fee_approvals": approvals,
        "errors": [],
        "warnings": warnings,
    }


@router.get("/schema")
async def upload_schema(user: dict = Depends(get_current_user)):
    """Column contract for the upload UI — labels, requiredness and allowed values."""
    _assert_can_upload(user)
    return {
        "row_limit": ROW_LIMIT,
        "max_bytes": MAX_UPLOAD_BYTES,
        "kinds": [
            {
                "kind": spec.kind,
                "sheet_name": spec.sheet_name,
                "title": spec.title,
                "columns": [
                    {
                        "label": c.label,
                        "required": c.required,
                        "allowed": list(c.allowed),
                        "help": c.help,
                        "example": c.example,
                    }
                    for c in spec.columns
                ],
            }
            for spec in SPECS.values()
        ],
    }


@router.get("/template")
async def download_template(
    user: dict = Depends(get_current_user),
    kind: Optional[str] = Query(None, description="Limit to one sheet: " + ", ".join(KINDS)),
    samples: bool = Query(True, description="Include filled sample rows"),
):
    """Multi-sheet XLSX template — also the accepted upload format."""
    _assert_can_upload(user)
    kinds = None
    if kind:
        if kind not in SPECS:
            raise HTTPException(400, f"kind must be one of: {', '.join(KINDS)}")
        kinds = [kind]
    data = await run_in_threadpool(build_workbook, kinds, include_samples=samples)
    fname = f"pws-alpha-bulk-upload{'-' + kind if kind else ''}.xlsx"
    return Response(
        content=data,
        media_type=XLSX_MEDIA,
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/template.csv", response_class=PlainTextResponse)
async def download_template_csv(
    user: dict = Depends(get_current_user),
    kind: str = Query("player", description="Sheet to export: " + ", ".join(KINDS)),
    samples: bool = Query(True),
):
    """Single-sheet CSV template, for tools that cannot open XLSX."""
    _assert_can_upload(user)
    if kind not in SPECS:
        raise HTTPException(400, f"kind must be one of: {', '.join(KINDS)}")
    return build_csv(kind, include_samples=samples)


async def _load_sheets(file: UploadFile, spec: Optional[SheetSpec]) -> Dict[str, List[Dict[str, str]]]:
    raw = await _read_capped(file)
    try:
        if spec is None:
            if raw[:2] != b"PK":
                raise HTTPException(
                    400,
                    "This looks like a CSV. A CSV holds one sheet, so pick a single "
                    "record type (Students, Players, Staff or Teachers) instead of "
                    "\"All sheets\" — or upload the .xlsx workbook.",
                )
            sheets = await run_in_threadpool(_parse_workbook, raw)
            if not sheets:
                raise HTTPException(
                    400,
                    "No recognised sheets found. Expected one or more of: "
                    + ", ".join(s.sheet_name for s in SPECS.values()),
                )
            return sheets
        rows = await run_in_threadpool(_parse_single_sheet, raw, file.filename or "", spec)
        return {spec.kind: rows}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Bulk upload parse failed for %s", file.filename)
        raise HTTPException(400, "Could not read the file. Upload the .xlsx template or a CSV export of it.")


@router.post("/workbook")
async def upload_workbook(
    request: Request,
    file: UploadFile = File(...),
    dry_run: bool = Query(False, description="Validate only; write nothing"),
    user: dict = Depends(get_current_user),
):
    """Ingest every recognised sheet in one workbook."""
    _assert_can_upload(user)
    _assert_declared_size(request)
    sheets = await _load_sheets(file, None)
    return await _ingest(sheets, user, dry_run=dry_run)


@router.post("/{kind}")
async def upload_kind(
    request: Request,
    kind: str,
    file: UploadFile = File(...),
    dry_run: bool = Query(False),
    user: dict = Depends(get_current_user),
):
    """Ingest a single-kind sheet. `kind` accepts the singular or plural form."""
    _assert_can_upload(user)
    _assert_declared_size(request)
    normalized = kind.rstrip("s") if kind.rstrip("s") in SPECS else kind
    spec = SPECS.get(normalized)
    if spec is None:
        raise HTTPException(404, f"Unknown upload kind '{kind}'. Use one of: {', '.join(KINDS)}")
    sheets = await _load_sheets(file, spec)
    return await _ingest(sheets, user, dry_run=dry_run)
