"""Build the bulk-upload workbook (and CSV) from the specs in bulk_ingest.

The generated file is also the accepted upload format, so a user can download,
fill in, and re-upload without touching the headers.
"""
from __future__ import annotations

import csv
import io
from typing import List, Optional, Sequence

from bulk_ingest import SPECS, SheetSpec

HEADER_FILL = "1E40AF"
REQUIRED_FILL = "DBEAFE"
HELP_FONT = "475569"


def _autosize(ws) -> None:
    for cells in ws.columns:
        try:
            width = max((len(str(c.value)) for c in cells if c.value is not None), default=10)
            ws.column_dimensions[cells[0].column_letter].width = min(max(width + 2, 12), 40)
        except Exception:
            pass


def _write_sheet(wb, spec: SheetSpec, *, include_samples: bool) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    ws = wb.create_sheet(spec.sheet_name)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    req_fill = PatternFill("solid", fgColor=REQUIRED_FILL)
    help_font = Font(italic=True, size=9, color=HELP_FONT)

    for idx, col in enumerate(spec.columns, start=1):
        cell = ws.cell(row=1, column=idx, value=col.label + (" *" if col.required else ""))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        help_cell = ws.cell(row=2, column=idx, value=col.help or ("Required" if col.required else ""))
        help_cell.font = help_font
        help_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        if col.required:
            help_cell.fill = req_fill

        if col.allowed:
            joined = ",".join(col.allowed)
            if len(joined) <= 250:
                letter = get_column_letter(idx)
                dv = DataValidation(type="list", formula1=f'"{joined}"', allow_blank=not col.required)
                dv.error = f"Pick a listed value for {col.label}"
                dv.errorTitle = "Invalid value"
                ws.add_data_validation(dv)
                dv.add(f"{letter}3:{letter}500")

    row = 3
    if include_samples:
        for sample in spec.samples:
            for idx, col in enumerate(spec.columns, start=1):
                ws.cell(row=row, column=idx, value=sample.get(col.label, ""))
            row += 1

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 34
    _autosize(ws)


def _write_readme(wb, kinds: Sequence[str], *, include_samples: bool) -> None:
    from openpyxl.styles import Alignment, Font

    ws = wb.create_sheet("Read Me", 0)
    lines: List[tuple] = [
        ("PWS & ALPHA Tracker — Bulk Upload Workbook", True),
        ("", False),
        ("How to use", True),
        ("1. Fill one row per person on the relevant sheet. Do not rename or reorder the header row.", False),
        ("2. Row 1 is the column name. Row 2 is guidance and is ignored on upload — leave it in place.", False),
        ("3. Columns marked * are required. Blank optional cells are simply skipped.", False),
        ("4. Delete the sample rows before uploading your own data." if include_samples
         else "4. Start entering data from row 3.", False),
        ("5. Upload the whole workbook — every sheet is ingested in one pass.", False),
        ("", False),
        ("Rules that are checked", True),
        ("• Dates accept 2026-04-15 or 15/04/2026.", False),
        ("• Mobile numbers must be 10 digits starting 6-9. +91 and leading 0 are stripped.", False),
        ("• Yes/No columns accept Yes, No, Y, N, true, false, 1, 0.", False),
        ("• Money columns must be whole rupees, not negative.", False),
        ("• Harding Park supports Daily players only.", False),
        ("• A student with Transport Enabled must have a Transport Distance.", False),
        ("• Duplicate name + date of birth is rejected, both within the file and against existing records.", False),
        ("• Enrollment IDs (APL - N) are allocated by the server; there is no column for them.", False),
        ("• A row whose name matches an existing record but has a different date of birth is accepted", False),
        ("  and reported as a warning, so genuine namesakes are not blocked.", False),
        ("", False),
        ("Nothing is saved unless every row passes validation — the upload is all-or-nothing,", False),
        ("so a failed upload leaves your data exactly as it was.", False),
        ("", False),
        ("Sheets in this workbook", True),
    ]
    for kind in kinds:
        spec = SPECS[kind]
        required = [c.label for c in spec.columns if c.required]
        lines.append((f"• {spec.sheet_name} — {spec.title}. Required: {', '.join(required)}", False))

    for i, (text, bold) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(bold=bold, size=13 if (bold and i == 1) else 11)
        cell.alignment = Alignment(vertical="top", wrap_text=False)
    ws.column_dimensions["A"].width = 110


def _write_reference(wb) -> None:
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet("Reference")
    ws.cell(row=1, column=1, value="Sheet").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=1, column=2, value="Column").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=1, column=3, value="Required").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=1, column=4, value="Accepted values / format").font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor=HEADER_FILL)
    for c in range(1, 5):
        ws.cell(row=1, column=c).fill = fill

    row = 2
    for spec in SPECS.values():
        for col in spec.columns:
            ws.cell(row=row, column=1, value=spec.sheet_name)
            ws.cell(row=row, column=2, value=col.label)
            ws.cell(row=row, column=3, value="Yes" if col.required else "")
            ws.cell(row=row, column=4, value=col.help or "Free text")
            row += 1
    ws.freeze_panes = "A2"
    _autosize(ws)


def build_workbook(kinds: Optional[Sequence[str]] = None, *, include_samples: bool = True) -> bytes:
    from openpyxl import Workbook

    selected = list(kinds) if kinds else list(SPECS.keys())
    wb = Workbook()
    wb.remove(wb.active)
    _write_readme(wb, selected, include_samples=include_samples)
    for kind in selected:
        _write_sheet(wb, SPECS[kind], include_samples=include_samples)
    _write_reference(wb)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_csv(kind: str, *, include_samples: bool = True) -> str:
    spec = SPECS[kind]
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow([c.label + (" *" if c.required else "") for c in spec.columns])
    if include_samples:
        for sample in spec.samples:
            writer.writerow([sample.get(c.label, "") for c in spec.columns])
    return buf.getvalue()
