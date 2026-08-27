"""Generate ready-to-upload seed sheets for bulk ingestion testing."""
import csv as _csv
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook

from bulk_ingest import PLAYER_SPEC, STAFF_SPEC, STUDENT_SPEC, TEACHER_SPEC

OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "docs")

STUDENTS = [
    {"Full Name": "Aarohi Sharma", "Admission Number": "SEED-STU-001", "Roll Number": "1",
     "Class": "Class VI", "Section": "A", "Gender": "Female", "Date of Birth": "2014-06-12",
     "Date of Admission": "2026-04-01", "Student Type": "Day School",
     "Father's Name": "Rakesh Sharma", "Mother's Name": "Sunita Sharma",
     "Emergency Contact": "9876500101", "Primary Mobile": "9876500101",
     "Email": "aarohi.sharma@example.com", "Address": "12 Boring Road",
     "Locality": "Boring Road", "City": "Patna"},
    {"Full Name": "Vivaan Kumar", "Admission Number": "SEED-STU-002", "Roll Number": "2",
     "Class": "Class IX", "Section": "B", "Gender": "Male", "Date of Birth": "2011-02-28",
     "Date of Admission": "15/04/2026", "Student Type": "Boarding",
     "Father's Name": "Manoj Kumar", "Emergency Contact": "9876500102",
     "Primary Mobile": "9876500102", "Hostel Resident": "Yes", "City": "Gaya"},
    {"Full Name": "Ishita Verma", "Admission Number": "SEED-STU-003",
     "Class": "Class III", "Section": "A", "Gender": "Female",
     "Date of Admission": "2026-04-05", "Student Type": "Day Boarding",
     "Primary Mobile": "9876500103", "Transport Enabled": "Yes",
     "Transport Distance": "Up to 5 km", "Transport Fee Monthly": "1200",
     "Father's Name": "Alok Verma", "City": "Patna"},
    {"Full Name": "Kabir Singh", "Admission Number": "SEED-STU-004",
     "Class": "Class X", "Section": "C", "Gender": "Male",
     "Date of Admission": "2026-04-02", "Student Type": "Day School",
     "Primary Mobile": "9876500104", "Fee Override: Tuition": "1500",
     "Fee Override: Exam Fee": "800", "Guardian Name": "Harpreet Singh", "City": "Patna"},
]

PLAYERS = [
    {"Full Name": "Rohan Mehta", "Mobile Number": "9876500201", "Centre": "Balua",
     "Sport": "Cricket", "Player Type": "Daily", "Slot": "Morning", "Skill Level": "Beginner",
     "Date of Admission": "2026-04-10", "Date of Birth": "2012-08-19", "Age": "13",
     "Father's Name": "Deepak Mehta", "Locality": "Kankarbagh", "City": "Patna"},
    {"Full Name": "Ananya Roy", "Mobile Number": "9876500202", "Centre": "Balua",
     "Sport": "Football", "Player Type": "Hostel", "Slot": "Evening",
     "Skill Level": "Intermediate", "Date of Admission": "10/04/2026",
     "Date of Birth": "2010-11-05", "Guardian Name": "Sujata Roy",
     "Guardian Phone": "9876500212", "City": "Patna"},
    {"Full Name": "Aditya Nair", "Mobile Number": "9876500203", "Centre": "Harding Park",
     "Sport": "Cricket", "Player Type": "Daily", "Slot": "Both", "Skill Level": "Advanced",
     "Date of Admission": "2026-04-12", "Monthly Fee Override": "9000", "City": "Patna"},
    {"Full Name": "Meera Joshi", "Mobile Number": "9876500204", "Centre": "Balua",
     "Sport": "Football", "Player Type": "Boarding", "Slot": "Morning",
     "Skill Level": "Beginner", "Date of Admission": "2026-04-15",
     "Boarding Class": "Class VIII", "Boarding Section": "A",
     "Transport Fee Monthly": "800", "City": "Patna"},
]

STAFF = [
    {"Full Name": "Sanjay Prasad", "Organization": "PWS", "Primary Mobile": "9876500301",
     "Employee ID": "SEED-EMP-001", "Department": "Administration",
     "Designation": "Office Assistant", "Gender": "Male",
     "Date of Joining": "2026-04-01", "City": "Patna"},
    {"Full Name": "Rekha Devi", "Organization": "ALPHA", "Primary Mobile": "9876500302",
     "Employee ID": "SEED-EMP-002", "Department": "Grounds", "Designation": "Groundskeeper",
     "Gender": "Female", "Centre": "Balua", "Date of Joining": "2026-04-03", "City": "Patna"},
    {"Full Name": "Imran Ali", "Organization": "BOTH", "Primary Mobile": "9876500303",
     "Employee ID": "SEED-EMP-003", "Department": "Transport", "Designation": "Driver",
     "Gender": "Male", "Date of Joining": "2026-04-05", "City": "Patna"},
]

TEACHERS = [
    {"Full Name": "Anjali Gupta", "Date of Birth": "1990-03-14",
     "Address": "44 Rajendra Nagar, Patna", "Mobile": "9876500401",
     "Personal Email": "anjali.gupta@example.com", "Aadhaar Number": "234567890123",
     "Qualification": "Master's Degree", "Last Job": "St Xavier High School",
     "Guardian Name": "Ramesh Gupta", "Guardian Mobile": "9876500411",
     "Reference Name": "Dr Sunil Jha", "Reference Mobile": "9876500421",
     "Teacher Designation": "CLASS_TEACHER", "Date of Joining": "2026-04-01"},
    {"Full Name": "Nikhil Ranjan", "Date of Birth": "1993-07-22",
     "Address": "8 Patliputra Colony, Patna", "Mobile": "9876500402",
     "Personal Email": "nikhil.ranjan@example.com", "Aadhaar Number": "345678901234",
     "Qualification": "B.Ed", "Last Job": "Delhi Public School",
     "Guardian Name": "Shashi Ranjan", "Guardian Mobile": "9876500412",
     "Reference Name": "Meera Nair", "Reference Mobile": "9876500422",
     "Teacher Designation": "TEACHER", "Date of Joining": "2026-04-01"},
]

BROKEN_PLAYERS = [
    {"Full Name": "Bad Mobile", "Mobile Number": "12345", "Centre": "Balua", "Sport": "Cricket",
     "Player Type": "Daily", "Slot": "Morning", "Skill Level": "Beginner",
     "Date of Admission": "2026-04-10"},
    {"Full Name": "Bad Centre", "Mobile Number": "9876500901", "Centre": "Nowhere",
     "Sport": "Cricket", "Player Type": "Daily", "Slot": "Morning", "Skill Level": "Beginner",
     "Date of Admission": "2026-04-10"},
    {"Full Name": "Harding Hostel Clash", "Mobile Number": "9876500902", "Centre": "Harding Park",
     "Sport": "Football", "Player Type": "Hostel", "Slot": "Evening", "Skill Level": "Beginner",
     "Date of Admission": "2026-04-10"},
    {"Full Name": "", "Mobile Number": "9876500903", "Centre": "Balua", "Sport": "Cricket",
     "Player Type": "Daily", "Slot": "Morning", "Skill Level": "Beginner",
     "Date of Admission": "2026-04-10"},
    {"Full Name": "Negative Money", "Mobile Number": "9876500904", "Centre": "Balua",
     "Sport": "Cricket", "Player Type": "Daily", "Slot": "Morning", "Skill Level": "Beginner",
     "Date of Admission": "2026-04-10", "Monthly Fee Override": "-500"},
]

SPECS = [
    (STUDENT_SPEC, "Students", STUDENTS, "students"),
    (PLAYER_SPEC, "Players", PLAYERS, "players"),
    (STAFF_SPEC, "Staff", STAFF, "staff"),
    (TEACHER_SPEC, "Teachers", TEACHERS, "teachers"),
]


def headers(spec):
    return [c.label + (" *" if c.required else "") for c in spec.columns]


def help_row(spec):
    return [c.help or (", ".join(c.allowed) if c.allowed else "") for c in spec.columns]


def row_for(spec, data):
    return [data.get(c.label, "") for c in spec.columns]


def add_sheet(wb, spec, sheet_name, rows):
    ws = wb.create_sheet(sheet_name)
    ws.append(headers(spec))
    ws.append(help_row(spec))
    for r in rows:
        ws.append(row_for(spec, r))
    for i, c in enumerate(spec.columns, start=1):
        letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[letter].width = max(14, min(len(c.label) + 4, 34))
    return ws


def build_all():
    os.makedirs(OUT, exist_ok=True)
    written = []

    wb = Workbook()
    wb.remove(wb.active)
    for spec, name, rows, _slug in SPECS:
        add_sheet(wb, spec, name, rows)
    combined = os.path.join(OUT, "seed-bulk-upload-all.xlsx")
    wb.save(combined)
    written.append(combined)

    for spec, name, rows, slug in SPECS:
        one = Workbook()
        one.remove(one.active)
        add_sheet(one, spec, name, rows)
        path = os.path.join(OUT, "seed-%s.xlsx" % slug)
        one.save(path)
        written.append(path)

        csv_path = os.path.join(OUT, "seed-%s.csv" % slug)
        with open(csv_path, "w", newline="", encoding="utf-8") as fh:
            w = _csv.writer(fh)
            w.writerow(headers(spec))
            w.writerow(help_row(spec))
            for r in rows:
                w.writerow(row_for(spec, r))
        written.append(csv_path)

    bad = Workbook()
    bad.remove(bad.active)
    add_sheet(bad, PLAYER_SPEC, "Players", BROKEN_PLAYERS)
    bad_path = os.path.join(OUT, "seed-players-with-errors.xlsx")
    bad.save(bad_path)
    written.append(bad_path)
    return written


if __name__ == "__main__":
    for p in build_all():
        print(os.path.basename(p), os.path.getsize(p), "bytes")
