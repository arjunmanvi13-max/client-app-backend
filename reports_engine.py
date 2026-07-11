"""Shared reporting filters, entity labels, and Excel/PDF export helpers."""
from __future__ import annotations

import io
from datetime import datetime
from typing import Optional, List, Dict, Any, Callable, Awaitable

from fastapi import HTTPException
from fastapi.responses import StreamingResponse

from core import (
    db,
    get_perm,
    is_super_admin,
    is_admin,
    now_utc,
    resolve_user_institution,
    person_entity_filter,
    fee_entity_filter,
    attendance_entity_filter,
    derive_person_entities,
    format_date_display,
    format_datetime_display,
    format_month_display,
)

REPORT_IDS = (
    "students",
    "players",
    "staff",
    "attendance-summary",
    "attendance-detail",
    "fee-collection",
    "outstanding-invoices",
    "payment-receipts",
    "marks-summary",
    "report-card-status",
)

ENTITY_LABELS = {"pws": "PWS", "alpha": "ALPHA", "both": "Combined"}


def _access_reports(user: dict) -> None:
    if is_super_admin(user) or is_admin(user):
        return
    if user.get("role") in ("principal", "vice_principal"):
        return
    if get_perm(user, "access_reports"):
        return
    raise HTTPException(403, "You do not have access to Reports.")


def resolve_entity(user: dict, entity: Optional[str]) -> str:
    """Map entity filter to PWS | ALPHA | BOTH."""
    raw = (entity or "").strip().lower()
    if raw in ("pws", "alpha", "both", "combined"):
        requested = "BOTH" if raw in ("both", "combined") else raw.upper()
    else:
        requested = None
    return resolve_user_institution(user, requested)


def entity_display(person_or_row: dict) -> str:
    """Human entity label for combined reports."""
    if person_or_row.get("entity_label"):
        return person_or_row["entity_label"]
    ent = person_or_row.get("entity_id")
    if ent:
        return ENTITY_LABELS.get(str(ent).lower(), str(ent).upper())
    ents = person_or_row.get("entities") or derive_person_entities(person_or_row)
    if len(ents) > 1 or person_or_row.get("organization") == "BOTH":
        return "PWS + ALPHA"
    if ents:
        return ents[0]
    org = (person_or_row.get("organization") or "").upper()
    return org or "—"


def _parse_date(d: Optional[str]) -> Optional[str]:
    if not d:
        return None
    return d[:10] if len(d) >= 10 else d


def build_meta(
    report_id: str,
    title: str,
    user: dict,
    entity: str,
    filters: dict,
    columns: List[str],
    rows: List[dict],
    summary: Optional[dict] = None,
    row_keys: Optional[List[str]] = None,
) -> dict:
    rk = row_keys or (list(rows[0].keys()) if rows else [])
    return {
        "report_id": report_id,
        "title": title,
        "entity_scope": entity,
        "entity_scope_label": ENTITY_LABELS.get(entity.lower(), entity) if entity != "BOTH" else "Combined",
        "filters_applied": filters,
        "columns": columns,
        "row_keys": rk,
        "rows": rows,
        "summary": summary or {"total_rows": len(rows)},
        "generated_at": now_utc().isoformat(),
        "generated_by": user.get("name"),
    }


def _subtitle(entity: str, filters: dict, user: dict) -> str:
    parts = [f"Entity: {ENTITY_LABELS.get(entity.lower(), entity)}"]
    for k in ("date_from", "date_to", "grade", "section", "sport", "centre", "status"):
        v = filters.get(k)
        if v:
            if k in ("date_from", "date_to"):
                parts.append(f"{k}: {format_date_display(v)}")
            else:
                parts.append(f"{k}: {v}")
    parts.append(f"Generated: {format_datetime_display(now_utc().isoformat())} by {user.get('name', '—')}")
    return " · ".join(parts)


def export_excel(title: str, columns: List[str], rows: List[List[Any]], subtitle: str, filename: str) -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columns), 1))
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=14, color="1E40AF")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(columns), 1))
    ws.cell(row=2, column=1, value=subtitle).font = Font(italic=True, color="475569", size=10)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E40AF")
    hr = 4
    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=hr, column=idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")
    for ri, row in enumerate(rows, start=hr + 1):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=val)
    for col_cells in ws.columns:
        try:
            max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 42)
        except Exception:
            pass
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def export_pdf(title: str, columns: List[str], rows: List[List[Any]], subtitle: str, filename: str) -> StreamingResponse:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas as pdfcanvas

    buf = io.BytesIO()
    c = pdfcanvas.Canvas(buf, pagesize=landscape(A4))
    W, H = landscape(A4)
    y = H - 20 * mm
    c.setFont("Helvetica-Bold", 14)
    c.drawString(20 * mm, y, title)
    y -= 8 * mm
    c.setFont("Helvetica", 8)
    c.drawString(20 * mm, y, subtitle[:120])
    y -= 10 * mm
    col_w = (W - 40 * mm) / max(len(columns), 1)
    c.setFont("Helvetica-Bold", 8)
    for i, col in enumerate(columns):
        c.drawString(20 * mm + i * col_w, y, str(col)[:18])
    y -= 6 * mm
    c.setFont("Helvetica", 7)
    for row in rows[:500]:
        if y < 15 * mm:
            c.showPage()
            y = H - 20 * mm
            c.setFont("Helvetica", 7)
        for i, val in enumerate(row):
            c.drawString(20 * mm + i * col_w, y, str(val)[:22] if val is not None else "")
        y -= 5 * mm
    c.save()
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


async def _section_label(section_id: Optional[str]) -> Optional[str]:
    if not section_id:
        return None
    sec = await db.sections.find_one({"id": section_id}, {"_id": 0, "label": 1})
    return sec.get("label") if sec else None


def _people_base_query(kind: str, entity: str, centre: Optional[str], sport: Optional[str], status: Optional[str], section_id: Optional[str], grade: Optional[str]) -> dict:
    q: dict = {"kind": kind}
    ent_f = person_entity_filter(entity)
    if ent_f:
        q = {"$and": [q, ent_f]}
    if centre and centre.lower() != "all":
        q["centre"] = centre
    if sport and sport.lower() != "all":
        q["sport"] = sport
    if status and status.lower() != "all":
        q["status"] = status
    if section_id:
        q["section_id"] = section_id
    elif grade and grade.lower() != "all":
        q["group"] = {"$regex": f"^{grade}"}
    return q


async def run_students(user: dict, entity: str, filters: dict) -> dict:
    q = _people_base_query("student", entity, filters.get("centre"), filters.get("sport"), filters.get("status"), filters.get("section_id"), filters.get("grade"))
    rows_raw = await db.people.find(q, {"_id": 0}).sort("name", 1).to_list(3000)
    columns = ["Entity", "Name", "Admission No.", "Roll No.", "Grade/Section", "Status", "Resident"]
    rows = []
    for p in rows_raw:
        rows.append({
            "entity_label": entity_display(p),
            "name": p.get("name"),
            "admission_number": p.get("admission_number"),
            "roll_number": p.get("roll_number"),
            "grade_section": p.get("group"),
            "status": p.get("status", "active"),
            "is_resident": "Yes" if p.get("is_resident") else "No",
        })
    keys = ["entity_label", "name", "admission_number", "roll_number", "grade_section", "status", "is_resident"]
    return build_meta("students", "Student List", user, entity, filters, columns, rows, row_keys=keys)


async def run_players(user: dict, entity: str, filters: dict) -> dict:
    q = _people_base_query("player", entity, filters.get("centre"), filters.get("sport"), filters.get("status"), None, None)
    rows_raw = await db.people.find(q, {"_id": 0}).sort("name", 1).to_list(3000)
    columns = ["Entity", "Name", "Player ID", "Centre", "Sport", "Category", "Slot", "Status"]
    rows = []
    for p in rows_raw:
        rows.append({
            "entity_label": entity_display(p),
            "name": p.get("name"),
            "player_id": p.get("player_id"),
            "centre": p.get("centre"),
            "sport": p.get("sport"),
            "category": p.get("player_type"),
            "slot": p.get("slot"),
            "status": p.get("status", "active"),
        })
    keys = ["entity_label", "name", "player_id", "centre", "sport", "category", "slot", "status"]
    return build_meta("players", "Player List", user, entity, filters, columns, rows, row_keys=keys)


async def run_staff(user: dict, entity: str, filters: dict) -> dict:
    q = _people_base_query("staff", entity, filters.get("centre"), filters.get("sport"), filters.get("status"), None, None)
    rows_raw = await db.people.find(q, {"_id": 0}).sort("name", 1).to_list(2000)
    columns = ["Entity", "Name", "Employee ID", "Role/Dept", "Centre", "Status"]
    rows = []
    for p in rows_raw:
        rows.append({
            "entity_label": entity_display(p),
            "name": p.get("name"),
            "employee_id": p.get("employee_id"),
            "role": p.get("group"),
            "centre": p.get("centre"),
            "status": p.get("status", "active"),
        })
    keys = ["entity_label", "name", "employee_id", "role", "centre", "status"]
    return build_meta("staff", "Staff List", user, entity, filters, columns, rows, row_keys=keys)


async def run_attendance_summary(user: dict, entity: str, filters: dict) -> dict:
    start = filters.get("date_from") or now_utc().strftime("%Y-%m-%d")
    end = filters.get("date_to") or start
    match: dict = {"date": {"$gte": start, "$lte": end}}
    if filters.get("sport") and filters["sport"].lower() != "all":
        match["sport"] = filters["sport"]
    if filters.get("centre") and filters["centre"].lower() != "all":
        match["centre"] = filters["centre"]
    if filters.get("section_id"):
        match["section_id"] = filters["section_id"]
    elif filters.get("grade") and filters["grade"].lower() != "all":
        match["group"] = {"$regex": f"^{filters['grade']}"}
    ent_f = attendance_entity_filter(entity)
    if ent_f:
        match = {"$and": [match, ent_f]}
    pipeline = [
        {"$match": match},
        {"$group": {
            "_id": {"entity_id": "$entity_id", "kind": "$kind", "status": "$status", "date": "$date"},
            "count": {"$sum": 1},
        }},
    ]
    agg = await db.attendance.aggregate(pipeline).to_list(2000)
    columns = ["Entity", "Date", "Kind", "Status", "Count"]
    rows = []
    totals = {"present": 0, "absent": 0, "late": 0, "leave": 0}
    for r in agg:
        st = r["_id"].get("status", "")
        totals[st] = totals.get(st, 0) + r["count"]
        ent = r["_id"].get("entity_id") or "—"
        rows.append({
            "entity_label": ENTITY_LABELS.get(str(ent).lower(), str(ent).upper()),
            "date": r["_id"].get("date"),
            "kind": r["_id"].get("kind"),
            "status": st,
            "count": r["count"],
        })
    rows.sort(key=lambda x: (x.get("date") or "", x.get("kind") or ""))
    keys = ["entity_label", "date", "kind", "status", "count"]
    return build_meta("attendance-summary", "Attendance Summary", user, entity, filters, columns, rows, {"totals_by_status": totals}, row_keys=keys)


async def run_attendance_detail(user: dict, entity: str, filters: dict) -> dict:
    start = filters.get("date_from") or now_utc().strftime("%Y-%m-%d")
    end = filters.get("date_to") or start
    match: dict = {"date": {"$gte": start, "$lte": end}}
    if filters.get("sport") and filters["sport"].lower() != "all":
        match["sport"] = filters["sport"]
    if filters.get("centre") and filters["centre"].lower() != "all":
        match["centre"] = filters["centre"]
    if filters.get("status") and filters["status"].lower() != "all":
        match["status"] = filters["status"]
    ent_f = attendance_entity_filter(entity)
    if ent_f:
        match = {"$and": [match, ent_f]}
    recs = await db.attendance.find(match, {"_id": 0}).sort([("date", -1), ("kind", 1)]).to_list(5000)
    pids = list({r["person_id"] for r in recs})
    people = await db.people.find({"id": {"$in": pids}}, {"_id": 0, "id": 1, "name": 1}).to_list(5000)
    names = {p["id"]: p["name"] for p in people}
    columns = ["Entity", "Date", "Name", "Kind", "Status", "Group", "Sport", "Centre", "Marked By"]
    rows = []
    for r in recs:
        ent = r.get("entity_id") or "—"
        rows.append({
            "entity_label": ENTITY_LABELS.get(str(ent).lower(), str(ent).upper()),
            "date": r.get("date"),
            "name": names.get(r.get("person_id"), r.get("person_id")),
            "kind": r.get("kind"),
            "status": r.get("status"),
            "group": r.get("group"),
            "sport": r.get("sport"),
            "centre": r.get("centre"),
            "marked_by": r.get("marked_by_name"),
        })
    keys = ["entity_label", "date", "name", "kind", "status", "group", "sport", "centre", "marked_by"]
    return build_meta("attendance-detail", "Attendance Detail", user, entity, filters, columns, rows, row_keys=keys)


async def run_fee_collection(user: dict, entity: str, filters: dict) -> dict:
    q: dict = {"status": "paid"}
    q.update(fee_entity_filter(entity))
    if filters.get("centre") and filters["centre"].lower() != "all":
        q["centre"] = filters["centre"]
    if filters.get("sport") and filters["sport"].lower() != "all":
        q["sport"] = filters["sport"]
    df, dt = filters.get("date_from"), filters.get("date_to")
    if df or dt:
        rng: dict = {}
        if df:
            rng["$gte"] = df
        if dt:
            rng["$lte"] = dt
        q["paid_at"] = rng
    fees = await db.fees.find(q, {"_id": 0}).sort("paid_at", -1).to_list(3000)
    columns = ["Entity", "Person", "Fee Type", "Amount", "Paid At", "Mode", "Centre", "Sport"]
    rows = []
    total = 0
    for f in fees:
        ent = f.get("entity_id") or ("pws" if f.get("organization") == "PWS" else "alpha")
        amt = int(f.get("amount_due") or 0)
        total += amt
        rows.append({
            "entity_label": ENTITY_LABELS.get(str(ent).lower(), str(ent).upper()),
            "person": f.get("person_name") or f.get("player_name"),
            "fee_type": f.get("fee_type"),
            "amount": amt,
            "paid_at": f.get("paid_at"),
            "mode": f.get("payment_mode"),
            "centre": f.get("centre"),
            "sport": f.get("sport"),
        })
    inv_q: dict = {}
    if entity == "PWS":
        inv_q["entity_id"] = "pws"
    elif entity == "ALPHA":
        inv_q["entity_id"] = "alpha"
    if df or dt:
        inv_q["created_at"] = q.get("paid_at", {})
    inv_payments = await db.payments.find(inv_q if inv_q else {}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for p in inv_payments:
        if df and (p.get("created_at") or "")[:10] < df:
            continue
        if dt and (p.get("created_at") or "")[:10] > dt:
            continue
        amt = int(p.get("amount") or 0)
        total += amt
        rows.append({
            "entity_label": ENTITY_LABELS.get(str(p.get("entity_id", "pws")).lower(), "PWS"),
            "person": p.get("person_name"),
            "fee_type": "Invoice payment",
            "amount": amt,
            "paid_at": p.get("created_at"),
            "mode": p.get("payment_mode"),
            "centre": "",
            "sport": "",
        })
    keys = ["entity_label", "person", "fee_type", "amount", "paid_at", "mode", "centre", "sport"]
    return build_meta("fee-collection", "Fee Collection", user, entity, filters, columns, rows, {"total_collected": total}, row_keys=keys)


async def run_outstanding_invoices(user: dict, entity: str, filters: dict) -> dict:
    q: dict = {"balance_due": {"$gt": 0}, "status": {"$nin": ["cancelled", "draft"]}}
    if entity == "PWS":
        q["entity_id"] = "pws"
    elif entity == "ALPHA":
        q["entity_id"] = "alpha"
    if filters.get("status") and filters["status"].lower() != "all":
        q["status"] = filters["status"]
    invs = await db.invoices.find(q, {"_id": 0}).sort("due_date", 1).to_list(2000)
    columns = ["Entity", "Invoice No.", "Person", "Status", "Total", "Paid", "Balance", "Due Date"]
    rows = []
    total_bal = 0
    for inv in invs:
        bal = int(inv.get("balance_due") or inv.get("outstanding_amount") or 0)
        total_bal += bal
        rows.append({
            "entity_label": ENTITY_LABELS.get(str(inv.get("entity_id", "pws")).lower(), "PWS"),
            "invoice_number": inv.get("invoice_number"),
            "person": inv.get("person_name"),
            "status": inv.get("status"),
            "total": int(inv.get("total_amount") or 0),
            "paid": int(inv.get("amount_paid") or 0),
            "balance": bal,
            "due_date": inv.get("due_date"),
        })
    keys = ["entity_label", "invoice_number", "person", "status", "total", "paid", "balance", "due_date"]
    return build_meta("outstanding-invoices", "Outstanding Invoices", user, entity, filters, columns, rows, {"total_outstanding": total_bal}, row_keys=keys)


async def run_payment_receipts(user: dict, entity: str, filters: dict) -> dict:
    q: dict = {}
    if entity == "PWS":
        q["entity_id"] = "pws"
    elif entity == "ALPHA":
        q["entity_id"] = "alpha"
    df, dt = filters.get("date_from"), filters.get("date_to")
    if df or dt:
        rng: dict = {}
        if df:
            rng["$gte"] = df
        if dt:
            rng["$lte"] = dt + "T23:59:59"
        q["created_at"] = rng
    payments = await db.payments.find(q, {"_id": 0}).sort("created_at", -1).to_list(3000)
    columns = ["Entity", "Receipt No.", "Invoice", "Person", "Amount", "Mode", "Date", "Collected By"]
    rows = []
    total = 0
    for p in payments:
        amt = int(p.get("amount") or 0)
        total += amt
        rows.append({
            "entity_label": ENTITY_LABELS.get(str(p.get("entity_id", "pws")).lower(), "PWS"),
            "receipt_number": p.get("receipt_number"),
            "invoice_id": p.get("invoice_id"),
            "person": p.get("person_name"),
            "amount": amt,
            "mode": p.get("payment_mode"),
            "date": (p.get("created_at") or "")[:10],
            "collected_by": p.get("collected_by_name"),
        })
    keys = ["entity_label", "receipt_number", "invoice_id", "person", "amount", "mode", "date", "collected_by"]
    return build_meta("payment-receipts", "Payment Receipts", user, entity, filters, columns, rows, {"total_amount": total}, row_keys=keys)


async def run_marks_summary(user: dict, entity: str, filters: dict) -> dict:
    q: dict = {}
    if entity == "PWS":
        q["entity_id"] = "pws"
    elif entity == "ALPHA":
        q["entity_id"] = "alpha"
    if filters.get("section_id"):
        q["section_id"] = filters["section_id"]
    if filters.get("status") and filters["status"].lower() != "all":
        q["status"] = filters["status"]
    marks = await db.academic_marks.find(q, {"_id": 0}).sort("entered_at", -1).to_list(3000)
    pids = list({m["person_id"] for m in marks})
    people = await db.people.find({"id": {"$in": pids}}, {"_id": 0, "id": 1, "name": 1, "group": 1}).to_list(3000)
    pmap = {p["id"]: p for p in people}
    sub_ids = list({m.get("subject_id") for m in marks if m.get("subject_id")})
    subs = await db.subjects.find({"id": {"$in": sub_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
    smap = {s["id"]: s["name"] for s in subs}
    columns = ["Entity", "Student", "Section", "Subject", "Marks", "Max", "Grade", "Status"]
    rows = []
    for m in marks:
        p = pmap.get(m["person_id"], {})
        rows.append({
            "entity_label": ENTITY_LABELS.get(str(m.get("entity_id", "pws")).lower(), "PWS"),
            "student": p.get("name"),
            "section": p.get("group"),
            "subject": smap.get(m.get("subject_id"), m.get("subject_id")),
            "marks": m.get("marks_obtained"),
            "max_marks": m.get("max_marks"),
            "grade": m.get("grade"),
            "status": m.get("status"),
        })
    keys = ["entity_label", "student", "section", "subject", "marks", "max_marks", "grade", "status"]
    return build_meta("marks-summary", "Marks Summary", user, entity, filters, columns, rows, row_keys=keys)


async def run_report_card_status(user: dict, entity: str, filters: dict) -> dict:
    q: dict = {}
    if entity == "PWS":
        q["entity_id"] = "pws"
    elif entity == "ALPHA":
        q["entity_id"] = "alpha"
    if filters.get("status") and filters["status"].lower() != "all":
        q["status"] = filters["status"]
    cards = await db.report_cards.find(q, {"_id": 0}).sort("updated_at", -1).to_list(2000)
    columns = ["Entity", "Student", "Section", "Term", "Status", "Published At", "Teacher Remark"]
    rows = []
    for c in cards:
        rows.append({
            "entity_label": ENTITY_LABELS.get(str(c.get("entity_id", "pws")).lower(), "PWS"),
            "student": c.get("person_name"),
            "section": c.get("section_label") or c.get("grade_section"),
            "term": c.get("exam_term_name"),
            "status": c.get("status"),
            "published_at": c.get("published_at"),
            "teacher_remark": (c.get("teacher_remark") or "")[:80],
        })
    by_status: dict = {}
    for c in cards:
        st = c.get("status", "draft")
        by_status[st] = by_status.get(st, 0) + 1
    keys = ["entity_label", "student", "section", "term", "status", "published_at", "teacher_remark"]
    return build_meta("report-card-status", "Report Card Status", user, entity, filters, columns, rows, {"by_status": by_status}, row_keys=keys)


RUNNERS: Dict[str, Callable[..., Awaitable[dict]]] = {
    "students": run_students,
    "players": run_players,
    "staff": run_staff,
    "attendance-summary": run_attendance_summary,
    "attendance-detail": run_attendance_detail,
    "fee-collection": run_fee_collection,
    "outstanding-invoices": run_outstanding_invoices,
    "payment-receipts": run_payment_receipts,
    "marks-summary": run_marks_summary,
    "report-card-status": run_report_card_status,
}


REPORT_CATALOG = [
    {"id": "students", "title": "Student List", "category": "People", "filters": ["entity", "grade", "section", "status"]},
    {"id": "players", "title": "Player List", "category": "People", "filters": ["entity", "centre", "sport", "status"]},
    {"id": "staff", "title": "Staff List", "category": "People", "filters": ["entity", "centre", "status"]},
    {"id": "attendance-summary", "title": "Attendance Summary", "category": "Attendance", "filters": ["entity", "date_range", "grade", "section", "centre", "sport"]},
    {"id": "attendance-detail", "title": "Attendance Detail", "category": "Attendance", "filters": ["entity", "date_range", "centre", "sport", "status"]},
    {"id": "fee-collection", "title": "Fee Collection", "category": "Finance", "filters": ["entity", "date_range", "centre", "sport"]},
    {"id": "outstanding-invoices", "title": "Outstanding Invoices", "category": "Finance", "filters": ["entity", "status"]},
    {"id": "payment-receipts", "title": "Payment Receipts", "category": "Finance", "filters": ["entity", "date_range"]},
    {"id": "marks-summary", "title": "Marks Summary", "category": "Academic", "filters": ["entity", "section", "status"]},
    {"id": "report-card-status", "title": "Report Card Status", "category": "Academic", "filters": ["entity", "status"]},
]


def _format_report_cell(key: str, value: Any) -> Any:
    if value is None or value == "":
        return value
    k = str(key).lower()
    if k.endswith("_at") or k in ("timestamp", "generated"):
        return format_datetime_display(str(value))
    if k.endswith("_date") or k == "date" or k.endswith("_month"):
        if k.endswith("_month"):
            return format_month_display(str(value))
        return format_date_display(str(value))
    if isinstance(value, str) and len(value) >= 10 and value[4:5] == "-" and value[7:8] == "-":
        return format_date_display(value)
    return value


def dict_rows_to_matrix(meta: dict) -> tuple[List[str], List[List[Any]]]:
    columns = meta["columns"]
    if not meta["rows"]:
        return columns, []
    keys = meta.get("row_keys") or list(meta["rows"][0].keys())
    matrix = [[_format_report_cell(k, r.get(k)) for k in keys] for r in meta["rows"]]
    return columns, matrix
