"""Bulk ingestion: template round-trip, validation, and all-or-nothing writes.

Requires a running backend (EXPO_PUBLIC_BACKEND_URL) with demo data seeded.
"""
import io
import os
import uuid

import pytest
import requests
from openpyxl import Workbook

BASE = (os.environ.get("EXPO_PUBLIC_BACKEND_URL") or "http://127.0.0.1:8000").rstrip("/") + "/api"
SUPER = ("superadmin@prarambhika.com", "Super@123")
XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

PLAYER_HEADERS = [
    "Full Name *", "Mobile Number *", "Centre *", "Sport *", "Player Type *",
    "Slot *", "Skill Level *", "Date of Admission *", "Date of Birth", "Monthly Fee Override",
]


@pytest.fixture(scope="module")
def headers():
    r = requests.post(f"{BASE}/auth/login", json={"email": SUPER[0], "password": SUPER[1]}, timeout=20)
    assert r.status_code == 200, r.text
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


def _player_sheet(rows):
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Players")
    ws.append(PLAYER_HEADERS)
    ws.append([""] * len(PLAYER_HEADERS))
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _row(name, *, mobile="9876500000", centre="Balua", sport="Cricket", ptype="Daily",
         slot="Morning", skill="Beginner", doa="2026-04-15", dob="", override=""):
    return [name, mobile, centre, sport, ptype, slot, skill, doa, dob, override]


def _upload(headers, data, *, path="players", dry_run=True):
    return requests.post(
        f"{BASE}/bulk-upload/{path}",
        headers=headers,
        params={"dry_run": str(dry_run).lower()},
        files={"file": ("upload.xlsx", data, XLSX)},
        timeout=60,
    )


class TestTemplate:
    def test_template_has_a_sheet_per_kind(self, headers):
        from openpyxl import load_workbook

        r = requests.get(f"{BASE}/bulk-upload/template", headers=headers, timeout=60)
        assert r.status_code == 200, r.text
        wb = load_workbook(io.BytesIO(r.content))
        for sheet in ("Students", "Players", "Staff", "Teachers", "Reference", "Read Me"):
            assert sheet in wb.sheetnames, wb.sheetnames

    def test_template_round_trips(self, headers):
        """The shipped template, uploaded unmodified, must validate cleanly."""
        r = requests.get(f"{BASE}/bulk-upload/template", headers=headers, timeout=60)
        assert r.status_code == 200
        up = requests.post(
            f"{BASE}/bulk-upload/workbook",
            headers=headers,
            params={"dry_run": "true"},
            files={"file": ("template.xlsx", r.content, XLSX)},
            timeout=60,
        )
        assert up.status_code == 200, up.text
        body = up.json()
        assert body["status"] == "ok", body
        assert body["counts"].get("student") and body["counts"].get("player")

    def test_csv_template_matches_columns(self, headers):
        r = requests.get(f"{BASE}/bulk-upload/template.csv", headers=headers,
                         params={"kind": "player"}, timeout=30)
        assert r.status_code == 200, r.text
        header_line = r.text.splitlines()[0]
        for required in ("Full Name", "Centre", "Sport", "Date of Admission"):
            assert required in header_line

    def test_csv_with_guidance_row_round_trips(self, headers):
        """Users save the XLSX template as CSV — its row-2 guidance must not parse as data."""
        r = requests.get(f"{BASE}/bulk-upload/template", headers=headers, timeout=60)
        assert r.status_code == 200

        from openpyxl import load_workbook

        ws = load_workbook(io.BytesIO(r.content))["Players"]
        rows = [
            [("" if c is None else str(c)) for c in row]
            for row in ws.iter_rows(values_only=True)
        ]
        import csv as _csv

        buf = io.StringIO()
        _csv.writer(buf).writerows(rows)
        csv_text = buf.getvalue()

        up = requests.post(
            f"{BASE}/bulk-upload/players",
            headers=headers,
            params={"dry_run": "true"},
            files={"file": ("players.csv", csv_text.encode("utf-8"), "text/csv")},
            timeout=60,
        )
        assert up.status_code == 200, up.text
        body = up.json()
        assert body["status"] == "ok", body

    def test_schema_lists_required_columns(self, headers):
        r = requests.get(f"{BASE}/bulk-upload/schema", headers=headers, timeout=30)
        assert r.status_code == 200, r.text
        kinds = {k["kind"]: k for k in r.json()["kinds"]}
        assert set(kinds) == {"student", "player", "staff", "teacher"}
        player_required = [c["label"] for c in kinds["player"]["columns"] if c["required"]]
        assert "Full Name" in player_required and "Centre" in player_required


class TestValidation:
    def test_bad_enum_is_rejected_with_row_number(self, headers):
        data = _player_sheet([_row("Bad Centre", centre="Nowhere")])
        body = _upload(headers, data).json()
        assert body["status"] == "validation_failed"
        assert body["errors"][0]["row"] == 3
        assert "Centre" in body["errors"][0]["errors"][0]

    def test_bad_mobile_is_rejected(self, headers):
        data = _player_sheet([_row("Bad Mobile", mobile="12345")])
        body = _upload(headers, data).json()
        assert body["status"] == "validation_failed"
        assert "Mobile" in body["errors"][0]["errors"][0]

    def test_missing_required_field_is_rejected(self, headers):
        data = _player_sheet([_row("")])
        body = _upload(headers, data).json()
        assert body["status"] == "validation_failed"
        assert "required" in body["errors"][0]["errors"][0].lower()

    def test_cross_field_rule_is_enforced(self, headers):
        data = _player_sheet([_row("Harding Hostel", centre="Harding Park", ptype="Hostel")])
        body = _upload(headers, data).json()
        assert body["status"] == "validation_failed"
        assert "Harding Park" in body["errors"][0]["errors"][0]

    def test_defense_colony_daily_only(self, headers):
        data = _player_sheet([_row("Defense Hostel", centre="Defense Colony", ptype="Hostel")])
        body = _upload(headers, data).json()
        assert body["status"] == "validation_failed"
        assert "Defense Colony" in body["errors"][0]["errors"][0]

    def test_negative_money_is_rejected(self, headers):
        data = _player_sheet([_row("Neg Money", override="-500")])
        body = _upload(headers, data).json()
        assert body["status"] == "validation_failed"
        assert "negative" in body["errors"][0]["errors"][0].lower()

    def test_in_file_duplicate_is_rejected(self, headers):
        data = _player_sheet([
            _row("Twin Row", mobile="9876500011", dob="2010-05-05"),
            _row("Twin Row", mobile="9876500012", dob="2010-05-05"),
        ])
        body = _upload(headers, data).json()
        assert body["status"] == "validation_failed"
        assert any("Duplicate of row" in e["errors"][0] for e in body["errors"])

    def test_flexible_date_formats_accepted(self, headers):
        data = _player_sheet([_row("Date Format", mobile="9876500013", doa="15/04/2026")])
        body = _upload(headers, data).json()
        assert body["status"] == "ok", body

    def test_unknown_column_is_reported(self, headers):
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("Players")
        ws.append(PLAYER_HEADERS + ["Favourite Colour"])
        ws.append([""] * (len(PLAYER_HEADERS) + 1))
        ws.append(_row("Extra Col", mobile="9876500014") + ["blue"])
        buf = io.BytesIO()
        wb.save(buf)
        body = _upload(headers, buf.getvalue()).json()
        assert body["status"] == "validation_failed"
        assert "Unrecognised column" in body["errors"][0]["errors"][-1]


class TestAllOrNothing:
    def test_failed_upload_writes_nothing(self, headers):
        marker = uuid.uuid4().hex[:8]
        good = f"Keeper {marker}"
        data = _player_sheet([
            _row(good, mobile="9876500021"),
            _row("Broken Row", mobile="9876500022", centre="Nowhere"),
        ])
        body = _upload(headers, data, dry_run=False).json()
        assert body["status"] == "validation_failed", body

        listing = requests.get(f"{BASE}/people", headers=headers,
                               params={"kind": "player", "q": good}, timeout=30)
        assert listing.status_code == 200, listing.text
        rows = listing.json()
        rows = rows if isinstance(rows, list) else rows.get("items", [])
        assert not [p for p in rows if p.get("name") == good], \
            "valid row was written despite the upload failing"

class TestImportAndRollback:
    """The write path — dry-run coverage alone never exercises _rollback."""

    def test_real_import_generates_fees_then_cleans_up(self, headers):
        marker = uuid.uuid4().hex[:6]
        name = f"QA Import {marker}"
        data = _player_sheet([_row(name, mobile=f"98761{marker[:5].translate(str.maketrans('abcdef', '123456'))}")])
        body = _upload(headers, data, dry_run=False).json()
        assert body["status"] == "ok", body
        assert body["counts"].get("player") == 1, body
        assert body.get("fees_created", 0) > 0, "imported players must get a fee schedule"

        listing = requests.get(f"{BASE}/people", headers=headers,
                               params={"kind": "player", "q": name}, timeout=30)
        rows = listing.json()
        rows = rows if isinstance(rows, list) else rows.get("data", rows.get("items", []))
        created = [p for p in rows if p.get("name") == name]
        assert created, "player was reported imported but is not in the roster"

        requests.delete(f"{BASE}/people/{created[0]['id']}", headers=headers, timeout=30)

    @pytest.mark.asyncio
    async def test_rollback_clears_people_fees_and_login_accounts(self, headers):
        """A write-time failure must leave no person, fee or staff login behind.

        Uniqueness clashes are caught during validation and never reach the write
        path, so this drives _rollback directly against the rows a real import
        creates: db.people, db.fees and the auto-provisioned staff login.
        """
        from routers.uploads import _rollback

        marker = uuid.uuid4().hex[:6]
        staff_name = f"QA Rollback Staff {marker}"
        player_name = f"QA Rollback Player {marker}"

        staff_headers = ["Full Name *", "Organization *", "Primary Mobile *", "Employee ID"]
        wb_staff = Workbook()
        wb_staff.remove(wb_staff.active)
        ws = wb_staff.create_sheet("Staff")
        ws.append(staff_headers)
        ws.append([""] * len(staff_headers))
        ws.append([staff_name, "PWS", "9876513001", f"QA-RB-{marker}"])
        buf = io.BytesIO()
        wb_staff.save(buf)

        staff_res = requests.post(
            f"{BASE}/bulk-upload/staff", headers=headers, params={"dry_run": "false"},
            files={"file": ("staff.xlsx", buf.getvalue(), XLSX)}, timeout=60,
        ).json()
        assert staff_res["status"] == "ok", staff_res

        player_res = _upload(headers, _player_sheet([_row(player_name, mobile="9876513002")]),
                             dry_run=False).json()
        assert player_res["status"] == "ok", player_res
        assert player_res.get("fees_created", 0) > 0, "player import generated no fees"

        from core import db

        if True:
            staff = await db.people.find_one({"name": staff_name}, {"_id": 0, "id": 1})
            player = await db.people.find_one({"name": player_name}, {"_id": 0, "id": 1})
            assert staff and player, "imported rows are missing before rollback"
            assert await db.users.count_documents({"person_id": staff["id"]}) == 1,                 "staff import did not create the login account this test is about"
            assert await db.fees.count_documents({"player_id": player["id"]}) > 0

            assert await _rollback("staff", [staff["id"]]) == []
            assert await _rollback("player", [player["id"]]) == []

            left = {
                "people": await db.people.count_documents(
                    {"id": {"$in": [staff["id"], player["id"]]}}),
                "logins": await db.users.count_documents({"person_id": staff["id"]}),
                "fees": await db.fees.count_documents({"player_id": player["id"]}),
            }
            assert left == {"people": 0, "logins": 0, "fees": 0},                 f"rollback left rows behind: {left}"
